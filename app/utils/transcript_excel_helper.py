# ============================================================
# Ported 1:1 from utils/transcript_excel_helper.py. Pure/sync (openpyxl,
# no DB access) — run via asyncio.to_thread() from the transcripts router.
#
# Interface change vs. Flask: takes (filename, payload: bytes) instead of
# a Werkzeug FileStorage, since the router reads the FastAPI UploadFile
# into bytes first.
# ============================================================

import re
from collections import Counter
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from app.utils.grade_points import get_grade_point
from app.utils.grading_helper import assign_relative_grades

SEMESTER_PATTERN = re.compile(r"^(?:trimester|semester)\s*(\d+)\s*$", re.IGNORECASE)
MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_STUDENTS = 5000
MAX_SEMESTERS = 20
MAX_SUBJECTS_PER_SEMESTER = 100


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _semester_number(value: Any) -> Optional[int]:
    match = SEMESTER_PATTERN.match(_display_value(value))
    return int(match.group(1)) if match else None


def _validate_file(filename: str, payload: bytes) -> BytesIO:
    if not filename:
        raise ValueError("Please choose an Excel workbook")
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("Transcript import supports .xlsx files only")
    if len(payload) > MAX_FILE_SIZE:
        raise ValueError("Excel workbook must be 10 MB or smaller")
    if not payload.startswith(b"PK"):
        raise ValueError("The selected file is not a valid .xlsx workbook")
    return BytesIO(payload)


def parse_transcript_workbook(filename: str, payload: bytes) -> Dict[str, Any]:
    stream = _validate_file(filename, payload)
    try:
        workbook = load_workbook(stream, data_only=True, read_only=False)
    except Exception as error:
        raise ValueError("Unable to read the transcript Excel workbook") from error

    semesters: List[Dict[str, Any]] = []
    seen_semesters = set()

    for worksheet in workbook.worksheets:
        row_number = 1
        while row_number <= worksheet.max_row:
            semester = _semester_number(worksheet.cell(row_number, 1).value)
            if semester is None:
                row_number += 1
                continue

            if semester in seen_semesters:
                raise ValueError(f"Semester {semester} appears more than once")
            if len(semesters) >= MAX_SEMESTERS:
                raise ValueError(f"A maximum of {MAX_SEMESTERS} semesters is supported")

            header_row = row_number + 1
            if header_row > worksheet.max_row:
                raise ValueError(f"Semester {semester} is missing its header row")

            first_header = _display_value(worksheet.cell(header_row, 1).value).lower()
            second_header = _display_value(worksheet.cell(header_row, 2).value).lower()
            if first_header not in {"course", "subject", "subject name", "course name"}:
                raise ValueError(f"Semester {semester} column A must be Course or Subject")
            if second_header not in {"credit", "credits", "subject credits"}:
                raise ValueError(f"Semester {semester} column B must be Credits")

            last_student_column = 2
            for column in range(3, worksheet.max_column + 1):
                if _display_value(worksheet.cell(header_row, column).value):
                    last_student_column = column

            student_ids = [
                _display_value(worksheet.cell(header_row, column).value)
                for column in range(3, last_student_column + 1)
            ]
            if not student_ids or any(not student_id for student_id in student_ids):
                raise ValueError(f"Semester {semester} contains an empty Student ID header")
            if len(student_ids) > MAX_STUDENTS:
                raise ValueError(f"A maximum of {MAX_STUDENTS} students is supported")

            duplicates = [student_id for student_id, count in Counter(student_ids).items() if count > 1]
            if duplicates:
                raise ValueError(
                    f"Semester {semester} contains duplicate Student IDs: " + ", ".join(duplicates[:5])
                )

            subjects = []
            seen_subjects = set()
            subject_row = header_row + 1
            while subject_row <= worksheet.max_row:
                subject_value = worksheet.cell(subject_row, 1).value
                if _semester_number(subject_value) is not None:
                    break
                subject_name = _display_value(subject_value)
                if not subject_name:
                    break
                if len(subjects) >= MAX_SUBJECTS_PER_SEMESTER:
                    raise ValueError(
                        f"Semester {semester} supports at most {MAX_SUBJECTS_PER_SEMESTER} subjects"
                    )

                subject_key = subject_name.casefold()
                if subject_key in seen_subjects:
                    raise ValueError(f"Semester {semester} contains duplicate subject {subject_name}")
                seen_subjects.add(subject_key)

                try:
                    credits = float(worksheet.cell(subject_row, 2).value)
                except (TypeError, ValueError) as error:
                    raise ValueError(
                        f"Credits must be numeric for {subject_name} in Semester {semester}"
                    ) from error
                if credits <= 0:
                    raise ValueError(f"Credits must be greater than zero for {subject_name}")

                marks = []
                for offset, column in enumerate(range(3, last_student_column + 1)):
                    value = worksheet.cell(subject_row, column).value
                    try:
                        numeric_mark = float(value)
                    except (TypeError, ValueError) as error:
                        raise ValueError(
                            f"Marks must be numeric for {student_ids[offset]}, {subject_name}, Semester {semester}"
                        ) from error
                    if not 0 <= numeric_mark <= 100:
                        raise ValueError(
                            f"Marks must be between 0 and 100 for {student_ids[offset]}, "
                            f"{subject_name}, Semester {semester}"
                        )
                    marks.append(numeric_mark)

                subjects.append({"subject": subject_name, "credits": credits, "marks": marks})
                subject_row += 1

            if not subjects:
                raise ValueError(f"Semester {semester} does not contain any subjects")

            semesters.append({"semester": semester, "student_ids": student_ids, "subjects": subjects})
            seen_semesters.add(semester)
            row_number = max(subject_row, row_number + 1)

    if not semesters:
        raise ValueError("No Semester or Trimester blocks were found in the workbook")

    semesters.sort(key=lambda item: item["semester"])
    expected_students = semesters[0]["student_ids"]
    for semester in semesters[1:]:
        if semester["student_ids"] != expected_students:
            raise ValueError("Student ID columns must be identical and in the same order for every semester")

    total_subjects = sum(len(semester["subjects"]) for semester in semesters)
    total_marks = len(expected_students) * total_subjects
    return {
        "student_ids": expected_students,
        "semesters": semesters,
        "summary": {
            "semester_count": len(semesters),
            "student_count": len(expected_students),
            "subject_count": total_subjects,
            "mark_count": total_marks,
            "semesters": [
                {
                    "semester": semester["semester"],
                    "subject_count": len(semester["subjects"]),
                    "total_credits": round(sum(subject["credits"] for subject in semester["subjects"]), 2),
                    "subjects": [
                        {"subject": subject["subject"], "credits": subject["credits"]}
                        for subject in semester["subjects"]
                    ],
                }
                for semester in semesters
            ],
        },
    }


def build_transcript_documents(
    parsed_workbook: Dict[str, Any],
    grading_config: List[Dict[str, Any]],
    institute_id: Any,
    batch_id: Any,
    import_id: str,
) -> List[Dict[str, Any]]:
    student_ids = parsed_workbook["student_ids"]
    cumulative = {student_id: {"credits": 0.0, "credit_points": 0.0} for student_id in student_ids}
    documents = []
    imported_at = datetime.now(timezone.utc)

    for semester in parsed_workbook["semesters"]:
        subject_results = {}
        for subject in semester["subjects"]:
            ranked_students = [
                {"student_id": student_id, "overall_total": subject["marks"][index]}
                for index, student_id in enumerate(student_ids)
            ]
            ranked_students.sort(key=lambda item: item["overall_total"], reverse=True)
            assign_relative_grades(ranked_students, grading_config)
            subject_results[subject["subject"]] = {
                item["student_id"]: {"marks": item["overall_total"], "grade": item["grade"]}
                for item in ranked_students
            }

        for student_id in student_ids:
            subjects = []
            semester_credits = 0.0
            semester_credit_points = 0.0
            semester_marks = 0.0

            for subject in semester["subjects"]:
                result = subject_results[subject["subject"]][student_id]
                grade_point = get_grade_point(result["grade"])
                credit_points = round(subject["credits"] * grade_point, 2)
                subjects.append({
                    "subject": subject["subject"],
                    "credits": subject["credits"],
                    "marks": result["marks"],
                    "grade": result["grade"],
                    "gradePoint": grade_point,
                    "creditPoints": credit_points,
                })
                semester_credits += subject["credits"]
                semester_credit_points += credit_points
                semester_marks += result["marks"]

            semester_credits = round(semester_credits, 2)
            semester_credit_points = round(semester_credit_points, 2)
            tgpa = round(semester_credit_points / semester_credits, 2) if semester_credits else 0
            cumulative[student_id]["credits"] += semester_credits
            cumulative[student_id]["credit_points"] += semester_credit_points
            cgpa = (
                round(cumulative[student_id]["credit_points"] / cumulative[student_id]["credits"], 2)
                if cumulative[student_id]["credits"] else 0
            )

            documents.append({
                "institute_id": institute_id,
                "batch_id": batch_id,
                "student_id": student_id,
                "semester_no": semester["semester"],
                "term_label": f"Semester {semester['semester']}",
                "subjects": subjects,
                "overall_total": round(semester_marks, 2),
                "total_credits": semester_credits,
                "total_credit_points": semester_credit_points,
                "tgpa": tgpa,
                "cgpa": cgpa,
                "import_id": import_id,
                "imported_at": imported_at,
            })

    return documents
