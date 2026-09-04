# ============================================================
# ACADEMIC TRANSCRIPTS ROUTER — Phase 5c
# Ported from controllers/institute/transcript_controller.py +
# routes/institute/transcript_routes.py (blueprint prefix "/transcript").
#
# Route ordering matters here: GET /transcript/imports and
# GET /transcript/{student_id} are both single-segment paths, so the
# literal "/imports" route MUST be registered before the catch-all
# "/{student_id}" route or it would never be reached (Starlette matches
# routes in registration order, same constraint as Flask's Werkzeug here).
#
# Response-shape note: /export/excel, /{student_id}/pdf and
# /{student_id}/certificate return a binary file on success but a JSON
# error body on failure — Flask did this by returning either a Response
# object or a (dict, status) tuple from the same function. FastAPI needs
# an explicit branch: JSONResponse for the error case, StreamingResponse
# for the binary case.
# ============================================================

import logging
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from bson import ObjectId
from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from motor.motor_asyncio import AsyncIOMotorDatabase
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.api.deps import get_current_identity
from app.db.mongodb import get_database
from app.models.relative_grading import build_grading_config
from app.models.transcript import get_imports, get_semesters, serialize_semester
from app.schemas.transcripts import TranscriptGenerateRequest
from app.utils.transcript_excel_helper import build_transcript_documents, parse_transcript_workbook
from app.utils.uploads import read_upload_capped
from app.utils.transcript_generation_helper import generate_transcript_for_semester

router = APIRouter(prefix="/transcript", dependencies=[Depends(get_current_identity)], tags=["transcripts"])

logger = logging.getLogger(__name__)


def _natural_key(value: Any) -> List[Any]:
    return [int(part) if part.isdigit() else part.casefold() for part in re.split(r"(\d+)", str(value))]


# ============================================================
# SHARED HELPERS
# ============================================================

async def _current_institute_id(identity: dict, db: AsyncIOMotorDatabase) -> ObjectId:
    user_id = identity.get("user_id")
    if not ObjectId.is_valid(user_id):
        raise HTTPException(status_code=401, detail="Invalid user id")

    user_object_id = ObjectId(user_id)
    user = await db["users"].find_one({"_id": user_object_id, "is_deleted": {"$ne": True}})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")

    institute = await db["instituteDetails"].find_one({"user_id": user_object_id, "is_deleted": {"$ne": True}})
    if institute:
        return institute["_id"]

    faculty = await db["facultyDetails"].find_one({"user_id": user_object_id, "is_deleted": {"$ne": True}})
    if faculty and faculty.get("institute_id"):
        return faculty["institute_id"]

    raise HTTPException(status_code=403, detail="Only institute administrators and faculty can access transcripts")


async def _owned_batch(batch_id: Optional[str], institute_id: ObjectId, db: AsyncIOMotorDatabase) -> Dict[str, Any]:
    if not batch_id or not ObjectId.is_valid(batch_id):
        raise HTTPException(status_code=400, detail="A valid batch is required")
    batch = await db["batchDetails"].find_one({
        "_id": ObjectId(batch_id), "institute_id": institute_id, "is_deleted": {"$ne": True},
    })
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found or access denied")
    return batch


async def _grading_config(institute_id: ObjectId, db: AsyncIOMotorDatabase) -> Optional[List[Dict[str, Any]]]:
    grading = await db["relativeGradings"].find_one({"university_id": institute_id})
    if not grading:
        return None
    return build_grading_config(grading)


async def _batch_metadata(batch: Dict[str, Any], db: AsyncIOMotorDatabase) -> Dict[str, str]:
    school = await db["schoolDetails"].find_one({"_id": batch.get("school_id")}) or {}
    programme = await db["programmeDetails"].find_one({"_id": batch.get("programme_id")}) or {}
    department: Dict[str, Any] = {}
    if batch.get("department_id"):
        department = await db["departmentDetails"].find_one({"_id": batch.get("department_id")}) or {}
    return {
        "school": school.get("school_name", ""),
        "programme": programme.get("programme_name", ""),
        "department": department.get("department_name", ""),
        "batch": batch.get("batch_name", ""),
    }


async def _student_name(student_id: str, institute_id: ObjectId, db: AsyncIOMotorDatabase) -> str:
    student = await db["studentDetails"].find_one({
        "institute_id": institute_id,
        "$or": [{"roll_number": student_id}, {"enrollment_number": student_id}],
        "is_deleted": {"$ne": True},
    })
    if not student or not student.get("user_id"):
        return student_id
    user = await db["users"].find_one({"_id": student["user_id"]}) or {}
    return user.get("fullName") or student_id


async def _get_academic_transcript(
    db: AsyncIOMotorDatabase, identity: dict, student_id: str, batch_id: Optional[str] = None
) -> Tuple[Dict[str, Any], int]:
    try:
        institute_id = await _current_institute_id(identity, db)

        batch = None
        if batch_id:
            batch = await _owned_batch(batch_id, institute_id, db)

        records = await get_semesters(db, student_id, institute_id, batch_id)
        if not records:
            return {"error": "No academic records found for this student"}, 404

        if batch is None:
            batch = await db["batchDetails"].find_one({
                "_id": records[0].get("batch_id"), "institute_id": institute_id,
            }) or {}
        metadata = await _batch_metadata(batch, db)
        semesters = [serialize_semester(record) for record in records]

        return {
            "status": "success",
            "studentName": await _student_name(str(student_id).strip(), institute_id, db),
            "studentId": str(student_id).strip(),
            **metadata,
            "semesters": semesters,
            "finalCGPA": semesters[-1]["cgpa"],
        }, 200
    except HTTPException as e:
        return {"error": e.detail}, e.status_code
    except Exception:
        logger.exception("Academic transcript fetch failed")
        return {"error": "Failed to fetch academic transcript"}, 500


# ============================================================
# IMPORT (EXCEL)
# ============================================================

@router.post("/import/preview")
async def preview_transcript_import(
    file: UploadFile = File(...),
    batch_id: str = Form(...),
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    try:
        institute_id = await _current_institute_id(identity, db)
        await _owned_batch(batch_id, institute_id, db)

        if not await _grading_config(institute_id, db):
            raise HTTPException(status_code=404, detail="Relative grading configuration not found")

        payload = await read_upload_capped(file)
        parsed = parse_transcript_workbook(file.filename or "", payload)
        return {
            "success": True,
            "message": "Workbook validated successfully",
            "summary": parsed["summary"],
            "student_id_sample": parsed["student_ids"][:10],
        }
    except HTTPException:
        raise
    except ValueError as error:
        return JSONResponse(status_code=400, content={"error": str(error)})
    except Exception:
        logger.exception("Transcript workbook preview failed")
        return JSONResponse(status_code=500, content={"error": "Unable to validate transcript workbook"})


@router.post("/import/confirm")
async def confirm_transcript_import(
    file: UploadFile = File(...),
    batch_id: str = Form(...),
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    import_id = str(uuid.uuid4())
    try:
        institute_id = await _current_institute_id(identity, db)
        batch = await _owned_batch(batch_id, institute_id, db)

        grading_config = await _grading_config(institute_id, db)
        if not grading_config:
            raise HTTPException(status_code=404, detail="Relative grading configuration not found")

        payload = await read_upload_capped(file)
        parsed = parse_transcript_workbook(file.filename or "", payload)
        documents = build_transcript_documents(parsed, grading_config, institute_id, batch["_id"], import_id)
        if not documents:
            raise HTTPException(status_code=400, detail="No transcript records were generated")

        await db["academic_transcripts"].insert_many(documents, ordered=True)
        await db["transcriptImports"].insert_one({
            "import_id": import_id,
            "institute_id": institute_id,
            "batch_id": batch["_id"],
            "filename": file.filename or "",
            "summary": parsed["summary"],
            "record_count": len(documents),
            "imported_by": ObjectId(identity["user_id"]),
            "imported_at": datetime.now(timezone.utc),
        })
        await db["academic_transcripts"].delete_many({
            "institute_id": institute_id, "batch_id": batch["_id"], "import_id": {"$ne": import_id},
        })

        return JSONResponse(status_code=201, content={
            "success": True,
            "message": "Transcript marks imported and calculated successfully",
            "import_id": import_id,
            "record_count": len(documents),
            "summary": parsed["summary"],
        })
    except HTTPException:
        await db["academic_transcripts"].delete_many({"import_id": import_id})
        await db["transcriptImports"].delete_many({"import_id": import_id})
        raise
    except ValueError as error:
        await db["academic_transcripts"].delete_many({"import_id": import_id})
        await db["transcriptImports"].delete_many({"import_id": import_id})
        return JSONResponse(status_code=400, content={"error": str(error)})
    except Exception:
        await db["academic_transcripts"].delete_many({"import_id": import_id})
        await db["transcriptImports"].delete_many({"import_id": import_id})
        logger.exception("Transcript workbook import failed")
        return JSONResponse(status_code=500, content={"error": "Unable to import and calculate transcript workbook"})


@router.get("/imports")
async def transcript_imports(
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    try:
        institute_id = await _current_institute_id(identity, db)
        imports = await get_imports(db, institute_id)
        return {
            "success": True,
            "imports": [
                {
                    **item,
                    "batch_id": str(item.get("batch_id", "")),
                    "institute_id": str(item.get("institute_id", "")),
                    "imported_by": str(item.get("imported_by", "")),
                    "imported_at": item.get("imported_at").isoformat() if item.get("imported_at") else None,
                }
                for item in imports
            ],
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Transcript import history fetch failed")
        return JSONResponse(status_code=500, content={"error": "Failed to fetch transcript imports"})


# ============================================================
# GENERATE (FROM LIVE EXAM DATA)
# ============================================================

@router.post("/generate/preview")
async def preview_transcript_generation(
    payload: TranscriptGenerateRequest,
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    try:
        institute_id = await _current_institute_id(identity, db)
        batch = await _owned_batch(payload.batch_id, institute_id, db)

        semester = payload.semester

        if not await _grading_config(institute_id, db):
            raise HTTPException(status_code=404, detail="Relative grading configuration not found")

        documents, summary = await generate_transcript_for_semester(db, institute_id, batch["_id"], semester)
        if not documents:
            raise HTTPException(status_code=404, detail="No exam results were found for this batch and semester")

        return {
            "success": True,
            "message": "Transcript generation validated successfully",
            "summary": summary,
        }
    except HTTPException:
        raise
    except Exception:
        logger.exception("Transcript generation preview failed")
        return JSONResponse(status_code=500, content={"error": "Unable to validate transcript generation"})


@router.post("/generate/confirm")
async def confirm_transcript_generation(
    payload: TranscriptGenerateRequest,
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    generation_id = str(uuid.uuid4())
    try:
        institute_id = await _current_institute_id(identity, db)
        batch = await _owned_batch(payload.batch_id, institute_id, db)

        semester = payload.semester

        if not await _grading_config(institute_id, db):
            raise HTTPException(status_code=404, detail="Relative grading configuration not found")

        documents, summary = await generate_transcript_for_semester(db, institute_id, batch["_id"], semester)
        if not documents:
            raise HTTPException(status_code=404, detail="No exam results were found for this batch and semester")

        for document in documents:
            document["import_id"] = generation_id

        await db["academic_transcripts"].insert_many(documents, ordered=True)
        await db["transcriptImports"].insert_one({
            "import_id": generation_id,
            "institute_id": institute_id,
            "batch_id": batch["_id"],
            "source": "generated",
            "summary": summary,
            "record_count": len(documents),
            "imported_by": ObjectId(identity["user_id"]),
            "imported_at": datetime.now(timezone.utc),
        })
        # Scoped to this semester only — unlike the Excel-import path's
        # delete_many (which deletes every semester for the batch), this
        # path only replaces a prior generation of the SAME semester.
        await db["academic_transcripts"].delete_many({
            "institute_id": institute_id, "batch_id": batch["_id"],
            "semester_no": int(semester), "import_id": {"$ne": generation_id},
        })

        return JSONResponse(status_code=201, content={
            "success": True,
            "message": "Transcript generated and calculated successfully",
            "import_id": generation_id,
            "record_count": len(documents),
            "summary": summary,
        })
    except HTTPException:
        await db["academic_transcripts"].delete_many({"import_id": generation_id})
        await db["transcriptImports"].delete_many({"import_id": generation_id})
        raise
    except Exception:
        await db["academic_transcripts"].delete_many({"import_id": generation_id})
        await db["transcriptImports"].delete_many({"import_id": generation_id})
        logger.exception("Transcript generation failed")
        return JSONResponse(status_code=500, content={"error": "Unable to generate and calculate transcript"})


# ============================================================
# EXCEL EXPORT
# ============================================================

@router.get("/export/excel")
async def export_transcript_excel(
    batch_id: str = Query(...),
    student_id: Optional[str] = Query(None),
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    try:
        institute_id = await _current_institute_id(identity, db)
        batch = await _owned_batch(batch_id, institute_id, db)

        normalized_student_id = str(student_id or "").strip()
        query: Dict[str, Any] = {"institute_id": institute_id, "batch_id": batch["_id"]}
        if normalized_student_id:
            query["student_id"] = {"$regex": f"^{re.escape(normalized_student_id)}$", "$options": "i"}

        records = [
            r async for r in
            db["academic_transcripts"].find(query, {"_id": 0}).sort([("student_id", 1), ("semester_no", 1)])
        ]
        if not records:
            message = (
                "No academic records found for this student" if normalized_student_id
                else "No academic transcript records found for this batch"
            )
            raise HTTPException(status_code=404, detail=message)

        grouped_records: Dict[str, List[Dict[str, Any]]] = {}
        for record in records:
            grouped_records.setdefault(record["student_id"], []).append(record)

        student_ids = sorted(grouped_records, key=_natural_key)
        student_names = {sid: await _student_name(sid, institute_id, db) for sid in student_ids}
        semester_numbers = sorted({record.get("semester_no") for record in records})
        metadata = await _batch_metadata(batch, db)

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Batch Summary"

        accent_fill = PatternFill("solid", fgColor="C7003D")
        dark_fill = PatternFill("solid", fgColor="172033")
        light_fill = PatternFill("solid", fgColor="FFF1F5")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(color="FFFFFF", bold=True, size=16)
        thin_side = Side(style="thin", color="E4E8EE")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        def apply_title(sheet, title, last_column):
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
            title_cell = sheet.cell(1, 1, title)
            title_cell.fill = dark_fill
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 28

        def apply_headers(sheet, row_number, headers):
            for column_number, header in enumerate(headers, start=1):
                cell = sheet.cell(row_number, column_number, header)
                cell.fill = accent_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = cell_border
            sheet.row_dimensions[row_number].height = 32

        def apply_data_rows(sheet, start_row, end_row, end_column):
            for row_number in range(start_row, end_row + 1):
                for column_number in range(1, end_column + 1):
                    cell = sheet.cell(row_number, column_number)
                    cell.border = cell_border
                    cell.alignment = Alignment(
                        vertical="center", horizontal="left" if column_number <= 2 else "center",
                    )
                    if row_number % 2 == 0:
                        cell.fill = light_fill

        def size_columns(sheet):
            for column_number in range(1, sheet.max_column + 1):
                values = [
                    str(sheet.cell(row_number, column_number).value or "")
                    for row_number in range(1, sheet.max_row + 1)
                ]
                width = min(max(max(map(len, values), default=0) + 2, 12), 32)
                sheet.column_dimensions[get_column_letter(column_number)].width = width

        summary_headers = ["Student ID", "Student Name"]
        for semester_number in semester_numbers:
            summary_headers.extend([f"Semester {semester_number} TGPA", f"Semester {semester_number} CGPA"])
        summary_headers.extend(["Final CGPA", "Total Credits", "Total Credit Points"])
        apply_title(summary_sheet, f"Academic Transcript Summary - {metadata['batch']}", len(summary_headers))
        summary_sheet.cell(2, 1, "Programme").font = Font(bold=True)
        summary_sheet.cell(2, 2, metadata["programme"])
        summary_sheet.cell(2, 3, "Department").font = Font(bold=True)
        summary_sheet.cell(2, 4, metadata["department"] or "—")
        summary_sheet.cell(2, 5, "Students").font = Font(bold=True)
        summary_sheet.cell(2, 6, len(student_ids))
        apply_headers(summary_sheet, 4, summary_headers)

        for row_number, current_student_id in enumerate(student_ids, start=5):
            student_records = grouped_records[current_student_id]
            records_by_semester = {record["semester_no"]: record for record in student_records}
            values: List[Any] = [current_student_id, student_names[current_student_id]]
            for semester_number in semester_numbers:
                semester_record = records_by_semester.get(semester_number, {})
                values.extend([semester_record.get("tgpa", ""), semester_record.get("cgpa", "")])
            final_record = max(student_records, key=lambda item: item.get("semester_no", 0))
            values.extend([
                final_record.get("cgpa", 0),
                round(sum(record.get("total_credits", 0) for record in student_records), 2),
                round(sum(record.get("total_credit_points", 0) for record in student_records), 2),
            ])
            for column_number, value in enumerate(values, start=1):
                summary_sheet.cell(row_number, column_number, value)

        apply_data_rows(summary_sheet, 5, 4 + len(student_ids), len(summary_headers))
        summary_sheet.freeze_panes = "C5"
        summary_sheet.auto_filter.ref = f"A4:{get_column_letter(len(summary_headers))}{4 + len(student_ids)}"
        size_columns(summary_sheet)

        for semester_number in semester_numbers:
            semester_records = [r for r in records if r.get("semester_no") == semester_number]
            semester_records.sort(key=lambda item: _natural_key(item["student_id"]))
            subject_names: List[str] = []
            for record in semester_records:
                for subject in record.get("subjects", []):
                    subject_name = subject.get("subject", "Subject")
                    if subject_name not in subject_names:
                        subject_names.append(subject_name)

            detail_headers = ["Student ID", "Student Name"]
            for subject_name in subject_names:
                detail_headers.extend([
                    f"{subject_name} Marks", f"{subject_name} Grade",
                    f"{subject_name} Credits", f"{subject_name} Credit Points",
                ])
            detail_headers.extend(["Overall Total", "Total Credits", "Total Credit Points", "TGPA", "CGPA"])

            detail_sheet = workbook.create_sheet(f"Semester {semester_number}")
            apply_title(detail_sheet, f"Semester {semester_number} Academic Results - {metadata['batch']}", len(detail_headers))
            apply_headers(detail_sheet, 3, detail_headers)

            for row_number, record in enumerate(semester_records, start=4):
                subjects_by_name = {s.get("subject", "Subject"): s for s in record.get("subjects", [])}
                values = [record["student_id"], student_names[record["student_id"]]]
                for subject_name in subject_names:
                    subject = subjects_by_name.get(subject_name, {})
                    values.extend([
                        subject.get("marks", ""), subject.get("grade", ""),
                        subject.get("credits", ""), subject.get("creditPoints", ""),
                    ])
                values.extend([
                    record.get("overall_total", 0), record.get("total_credits", 0),
                    record.get("total_credit_points", 0), record.get("tgpa", 0), record.get("cgpa", 0),
                ])
                for column_number, value in enumerate(values, start=1):
                    detail_sheet.cell(row_number, column_number, value)

            apply_data_rows(detail_sheet, 4, 3 + len(semester_records), len(detail_headers))
            detail_sheet.freeze_panes = "C4"
            detail_sheet.auto_filter.ref = f"A3:{get_column_letter(len(detail_headers))}{3 + len(semester_records)}"
            size_columns(detail_sheet)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        safe_batch = re.sub(r"[^A-Za-z0-9_-]+", "_", metadata["batch"] or "batch")
        if normalized_student_id:
            safe_student_id = re.sub(r"[^A-Za-z0-9_-]+", "_", records[0]["student_id"])
            filename = f"academic_transcript_{safe_student_id}.xlsx"
        else:
            filename = f"batch_academic_transcripts_{safe_batch}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException as e:
        return JSONResponse(status_code=e.status_code, content={"error": e.detail})
    except Exception:
        logger.exception("Academic transcript Excel export failed")
        return JSONResponse(status_code=500, content={"error": "Failed to export academic transcript Excel"})


# ============================================================
# PDF / CERTIFICATE
# ============================================================

@router.get("/{student_id}/pdf")
async def download_transcript_pdf(
    student_id: str,
    batch_id: Optional[str] = Query(None),
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    result, status = await _get_academic_transcript(db, identity, student_id, batch_id)
    if status != 200:
        return JSONResponse(status_code=status, content=result)

    output = BytesIO()
    document = SimpleDocTemplate(
        output, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TranscriptTitle", parent=styles["Title"], alignment=TA_CENTER, fontSize=18,
        textColor=colors.HexColor("#172033"),
    )
    story: List[Any] = [Paragraph("Academic Transcript", title_style), Spacer(1, 5 * mm)]
    metadata_rows = [
        ["Student Name", result["studentName"], "Student ID", result["studentId"]],
        ["Programme", result["programme"], "Department", result["department"]],
        ["Batch", result["batch"], "Final CGPA", result["finalCGPA"]],
    ]
    metadata_table = Table(metadata_rows, colWidths=[28 * mm, 55 * mm, 28 * mm, 55 * mm])
    metadata_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.extend([metadata_table, Spacer(1, 6 * mm)])

    co_text_style = ParagraphStyle(
        "COBreakdown", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#4B5563"),
    )

    for semester in result["semesters"]:
        story.append(Paragraph(f"Semester {semester['semester']}", styles["Heading2"]))
        rows: List[Any] = [["Subject", "Credits", "Marks", "Grade", "GP", "Credit Points"]]
        co_row_indices: List[int] = []  # rows that need a full-width SPAN + smaller font

        for subject in semester["subjects"]:
            rows.append([
                subject.get("subject", ""), subject.get("credits", 0), subject.get("marks", 0),
                subject.get("grade", ""), subject.get("gradePoint", 0), subject.get("creditPoints", 0),
            ])

            co_breakdown = subject.get("coBreakdown") or []
            if co_breakdown:
                co_text = "  |  ".join(
                    f"{co['co_code']}: {co['obtained_marks']}/{co['max_marks']} ({co['percentage']}%)"
                    for co in co_breakdown
                )
                co_row_indices.append(len(rows))
                rows.append([Paragraph(co_text, co_text_style), "", "", "", "", ""])

        rows.append([
            "Semester Total", semester["totalCredits"], semester["overallTotal"], "",
            f"TGPA {semester['tgpa']}", semester["totalCreditPoints"],
        ])
        semester_table = Table(rows, repeatRows=1, colWidths=[58 * mm, 20 * mm, 22 * mm, 20 * mm, 22 * mm, 30 * mm])
        table_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F3F4F6")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]
        for ri in co_row_indices:
            table_style.append(("SPAN", (0, ri), (-1, ri)))
            table_style.append(("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#F9FAFB")))
            table_style.append(("TOPPADDING", (0, ri), (-1, ri), 1))
            table_style.append(("BOTTOMPADDING", (0, ri), (-1, ri), 1))
        semester_table.setStyle(TableStyle(table_style))
        story.extend([semester_table, Spacer(1, 5 * mm)])

    document.build(story)
    output.seek(0)
    safe_student_id = re.sub(r"[^A-Za-z0-9_-]+", "_", str(student_id))
    return StreamingResponse(
        output, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=academic_transcript_{safe_student_id}.pdf"},
    )


@router.get("/{student_id}/certificate")
async def download_completion_certificate_pdf(
    student_id: str,
    batch_id: Optional[str] = Query(None),
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    """
    v1 eligibility: a transcript row exists for every semester of the
    programme (batch.total_semesters). No CGPA/pass-rate check in v1.
    """
    result, status = await _get_academic_transcript(db, identity, student_id, batch_id)
    if status != 200:
        return JSONResponse(status_code=status, content=result)

    institute_id = await _current_institute_id(identity, db)

    if not batch_id or not ObjectId.is_valid(batch_id):
        return JSONResponse(status_code=400, content={"error": "A valid batch is required"})
    batch = await _owned_batch(batch_id, institute_id, db)

    total_semesters = batch.get("total_semesters", 0) or 0
    if total_semesters <= 0 or len(result["semesters"]) < total_semesters:
        return JSONResponse(status_code=400, content={"error": "Student has not completed all semesters of the programme"})

    total_credits = round(sum(s.get("totalCredits", 0) for s in result["semesters"]), 2)

    output = BytesIO()
    document = SimpleDocTemplate(
        output, pagesize=A4, rightMargin=20 * mm, leftMargin=20 * mm, topMargin=25 * mm, bottomMargin=25 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CertificateTitle", parent=styles["Title"], alignment=TA_CENTER, fontSize=20,
        textColor=colors.HexColor("#172033"),
    )
    subtitle_style = ParagraphStyle(
        "CertificateSubtitle", parent=styles["Normal"], alignment=TA_CENTER, fontSize=12,
        textColor=colors.HexColor("#4B5563"),
    )
    body_style = ParagraphStyle(
        "CertificateBody", parent=styles["Normal"], alignment=TA_CENTER, fontSize=11, leading=17,
        spaceBefore=8 * mm, spaceAfter=8 * mm,
    )

    story: List[Any] = [
        Paragraph(result.get("school", ""), subtitle_style),
        Spacer(1, 4 * mm),
        Paragraph("Certificate of Completion", title_style),
        Spacer(1, 10 * mm),
        Paragraph(
            f"This is to certify that <b>{result['studentName']}</b> "
            f"(Student ID {result['studentId']}), a student of "
            f"<b>{result['programme']}</b> at {result.get('school', '')}, "
            f"has successfully completed all requirements of the "
            f"{total_semesters}-semester programme in the "
            f"<b>{result['batch']}</b> batch and has been awarded a final "
            f"CGPA of <b>{result['finalCGPA']}</b> out of 10.",
            body_style,
        ),
    ]

    summary_rows = [
        ["Programme", result["programme"], "Department", result["department"]],
        ["Batch", result["batch"], "Total Credits", total_credits],
        ["Final CGPA", result["finalCGPA"], "", ""],
    ]
    summary_table = Table(summary_rows, colWidths=[32 * mm, 55 * mm, 32 * mm, 55 * mm])
    summary_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.extend([Spacer(1, 6 * mm), summary_table, Spacer(1, 20 * mm)])

    issue_date = datetime.now(timezone.utc).strftime("%d %B %Y")
    footer_rows = [
        [f"Issued on {issue_date}", "", ""],
        ["", "", ""],
        ["_____________________", "", "_____________________"],
        ["Controller of Examinations", "", "Registrar"],
    ]
    footer_table = Table(footer_rows, colWidths=[58 * mm, 40 * mm, 58 * mm])
    footer_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (2, -1), "CENTER"),
    ]))
    story.append(footer_table)

    document.build(story)
    output.seek(0)
    safe_student_id = re.sub(r"[^A-Za-z0-9_-]+", "_", str(student_id))
    return StreamingResponse(
        output, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=certificate_{safe_student_id}.pdf"},
    )


# ============================================================
# JSON READ ENDPOINTS
# ============================================================

@router.get("/{student_id}/semester/{semester_no}")
async def get_semester_transcript(
    student_id: str,
    semester_no: int,
    batch_id: Optional[str] = Query(None),
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    result, status = await _get_academic_transcript(db, identity, student_id, batch_id)
    if status != 200:
        return JSONResponse(status_code=status, content=result)

    semester = next((item for item in result["semesters"] if item["semester"] == semester_no), None)
    if not semester:
        return JSONResponse(status_code=404, content={"error": f"Semester {semester_no} was not found"})

    return {"status": "success", "studentId": result["studentId"], "semester": semester}


@router.get("/{student_id}")
async def get_academic_transcript_route(
    student_id: str,
    batch_id: Optional[str] = Query(None),
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    result, status = await _get_academic_transcript(db, identity, student_id, batch_id)
    if status != 200:
        return JSONResponse(status_code=status, content=result)
    return result
