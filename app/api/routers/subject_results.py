# ============================================================
# SUBJECT / COMBINED RESULTS ROUTER
# Ported from controllers/institute/subject_controller.py's results
# functions + controllers/institute/excel_controller.py.
# ============================================================

import asyncio
import html as html_module
import io
import re
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorDatabase

from app.api.deps import (
    get_current_identity,
    get_current_user_and_faculty_details,
    get_current_user_and_institute,
    resolve_current_institute_id,
    validate_entity_ownership,
)
from app.db.mongodb import get_database
from app.models.relative_grading import build_grading_config
from app.services.co_excel import CoReportError, _normalise_answer, build_co_report_workbook, fetch_co_report_data
from app.utils.composite_helper import compute_weighted_composite
from app.utils.grade_points import get_grade_point
from app.utils.grading_helper import assign_relative_grades

router = APIRouter(dependencies=[Depends(get_current_identity)], tags=["subject-results"])


# ============================================================
# HELPERS
# ============================================================

def _extract_student_id(filename: str) -> str:
    """
    Extract student roll number from filename.
    e.g. "ME-304-021.pdf" -> "021", "CS101-A42.pdf" -> "A42"
    Falls back to the full stem when no hyphen is present.
    """
    stem = filename.rsplit(".", 1)[0]
    parts = stem.rsplit("-", 1)
    return parts[-1].strip() if parts else stem.strip()


def _natural_key(value: Any):
    """Sort key so "Student 2" comes before "Student 10"."""
    return [int(part) if part.isdigit() else part.casefold() for part in re.split(r"(\d+)", str(value))]


# ============================================================
# SUBJECT RESULTS
# ============================================================

async def _get_subject_results(
    db: AsyncIOMotorDatabase, subject_id: str, institute_id: ObjectId
) -> Tuple[Dict[str, Any], int]:
    if not ObjectId.is_valid(subject_id):
        return {"error": "Invalid subject_id"}, 400

    subject = await db["subjectDetails"].find_one({
        "_id": ObjectId(subject_id), "institute_id": institute_id, "is_deleted": {"$ne": True},
    })
    if not subject:
        return {"error": "Subject not found or unauthorized"}, 404

    subject_meta = {
        "id": str(subject["_id"]),
        "subject_name": subject.get("subject_name"),
        "subject_code": subject.get("subject_code"),
        "credits": subject.get("credits", 0),
    }

    exams = [e async for e in db["newsavedDocs"].find({"subject_id": ObjectId(subject_id)}).sort("exam_date", 1)]
    if not exams:
        return {
            "subject": subject_meta, "exams": [], "students": [],
            "total_students": 0, "total_exams": 0,
        }, 200

    exam_oids = [e["_id"] for e in exams]
    exam_id_str = [str(e["_id"]) for e in exams]

    exams_meta = []
    for e in exams:
        exam_date = e.get("exam_date")
        exams_meta.append({
            "exam_id": str(e["_id"]),
            "exam_title": e.get("exam_title", ""),
            "exam_type": e.get("exam_type", ""),
            "exam_date": exam_date.strftime("%Y-%m-%d") if exam_date else None,
            "weightage_percent": e.get("weightage") or 0,
            "folder_name": e.get("folder_name", ""),
        })
    weightage_by_exam = {em["exam_id"]: em["weightage_percent"] for em in exams_meta}

    answers = [a async for a in db["answerDetails"].find({"exam_id": {"$in": exam_oids}})]

    student_map: Dict[str, Dict[str, Any]] = {}
    # Per-student, per-CO running totals across every exam for this subject —
    # feeds the transcript's per-subject CO breakdown (marks + percentage per
    # CO). Reuses co_excel.py's _normalise_answer so the numbers here stay
    # consistent with the CO-Excel report's own CO aggregation.
    student_co_totals: Dict[str, Dict[str, Dict[str, float]]] = {}
    for ans in answers:
        eid = str(ans.get("exam_id", ""))
        filename = ans.get("filename", "")
        sid = _extract_student_id(filename)
        student_map.setdefault(sid, {})

        final_marks = ans.get("total_final_marks") or ans.get("total_ai_marks") or 0
        max_marks = ans.get("total_max_marks") or 0
        percentage = round(final_marks / max_marks * 100, 2) if max_marks else 0.0

        student_map[sid][eid] = {
            "answer_id": str(ans["_id"]),
            "final_marks": final_marks,
            "max_marks": max_marks,
            "percentage": percentage,
            "filename": filename,
            "evaluated_report_url": ans.get("evaluated_report_url"),
            "reviewed_by_professor": ans.get("reviewed_by_professor", False),
        }

        co_totals = student_co_totals.setdefault(sid, {})
        for q in _normalise_answer(ans)["questions"]:
            for co in q["cos"]:
                bucket = co_totals.setdefault(co["co_code"], {"obtained": 0.0, "max": 0.0})
                bucket["obtained"] += co["obtained_marks"]
                bucket["max"] += co["max_marks"]

    students_list = []
    for sid in sorted(student_map.keys(), key=_natural_key):
        results = {eid: student_map[sid].get(eid) for eid in exam_id_str}
        co_summary = [
            {
                "co_code": code,
                "obtained_marks": round(totals["obtained"], 2),
                "max_marks": round(totals["max"], 2),
                "percentage": round(totals["obtained"] / totals["max"] * 100, 2) if totals["max"] else 0.0,
            }
            for code, totals in sorted(student_co_totals.get(sid, {}).items())
        ]
        students_list.append({"student_id": sid, "results": results, "co_summary": co_summary})

    for student in students_list:
        exam_results = [
            {
                "exam_id": eid,
                "weightage": weightage_by_exam.get(eid, 0),
                "percentage": (student["results"][eid]["percentage"] if student["results"].get(eid) else None),
            }
            for eid in exam_id_str
        ]
        student.update(compute_weighted_composite(exam_results))

    grading = await db["relativeGradings"].find_one({"university_id": institute_id})
    if grading:
        grading_config = build_grading_config(grading)
        gradable_students = [s for s in students_list if s.get("composite_percentage") is not None]
        gradable_students.sort(key=lambda s: s["composite_percentage"], reverse=True)
        if gradable_students:
            graded = [{**s, "overall_total": s["composite_percentage"]} for s in gradable_students]
            assign_relative_grades(graded, grading_config, max_marks=100)
            for student, result in zip(gradable_students, graded):
                student["course_grade"] = result.get("grade")
                student["grade_point"] = get_grade_point(result.get("grade"))

    return {
        "subject": subject_meta,
        "exams": exams_meta,
        "students": students_list,
        "total_students": len(students_list),
        "total_exams": len(exams_meta),
    }, 200


@router.get("/subject/result/{subject_id}")
async def get_subject_result(
    subject_id: str, identity: dict = Depends(get_current_identity), db: AsyncIOMotorDatabase = Depends(get_database)
):
    user, institute_id, error = await get_current_user_and_institute(identity, db)
    if error:
        message, code = error
        raise HTTPException(status_code=code, detail=message)

    body, code = await _get_subject_results(db, subject_id, institute_id)
    if code >= 400:
        raise HTTPException(status_code=code, detail=body.get("error"))
    return body


@router.get("/faculty/subject/result/{subject_id}")
async def get_faculty_subject_result(
    subject_id: str, identity: dict = Depends(get_current_identity), db: AsyncIOMotorDatabase = Depends(get_database)
):
    user, faculty_id, error = await get_current_user_and_faculty_details(identity, db)
    if error:
        message, code = error
        raise HTTPException(status_code=code, detail=message)

    if not ObjectId.is_valid(subject_id):
        raise HTTPException(status_code=400, detail="Invalid subject_id")

    subject = await db["subjectDetails"].find_one({"_id": ObjectId(subject_id), "is_deleted": {"$ne": True}})
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")

    if subject.get("faculty_id") != faculty_id:
        raise HTTPException(status_code=403, detail="You are not assigned to this subject")

    body, code = await _get_subject_results(db, subject_id, subject["institute_id"])
    if code >= 400:
        raise HTTPException(status_code=code, detail=body.get("error"))
    return body


# ============================================================
# COMBINED RESULTS
# ============================================================

async def _get_combined_results(
    db: AsyncIOMotorDatabase, institute_id: ObjectId, batch_id: str, semester: str
) -> Tuple[Dict[str, Any], int]:
    if not ObjectId.is_valid(batch_id):
        return {"error": "Invalid batch_id"}, 400
    if not str(semester).isdigit():
        return {"error": "Invalid semester"}, 400

    grading = await db["relativeGradings"].find_one({"university_id": institute_id})
    if not grading:
        return {"error": "Relative grading configuration not found"}, 404

    grading_config = build_grading_config(grading)

    # ── Source 1: imported Excel marks ──────────────────────────────────
    imported_docs = [
        d async for d in db["importedMarks"].find({
            "institute_id": institute_id, "batch_id": ObjectId(batch_id), "semester": int(semester),
        })
    ]

    if imported_docs:
        combined_result = [
            {"student_id": doc.get("student_name"), "overall_total": doc.get("marks", 0)}
            for doc in imported_docs
        ]
        combined_result.sort(key=lambda x: x["overall_total"], reverse=True)
        for index, student in enumerate(combined_result, start=1):
            student["rank"] = index

        combined_result = assign_relative_grades(combined_result, grading_config)

        # This view only has this one semester's imported marks to work from —
        # no verified prior-semester history is available here (that's what
        # the Academic Transcript feature is for). Report each student's own
        # single-semester GPA as their CGPA rather than a shared placeholder
        # figure; a real cumulative CGPA belongs to the Transcript feature.
        total_credits = 15
        for student in combined_result:
            grade_point = get_grade_point(student.get("grade"))
            student["total_credits"] = total_credits
            student["total_credit_points"] = round(grade_point * total_credits, 2)
            student["tgpa"] = grade_point
            student["cgpa"] = grade_point

        return {"success": True, "data": combined_result}, 200

    # ── Source 2: exam-based aggregation, per-subject grading ──────────
    subjects = [
        s async for s in db["subjectDetails"].find({
            "batch_id": ObjectId(batch_id), "semester": int(semester),
            "institute_id": institute_id, "is_deleted": False,
        })
    ]

    total_semester_credits = sum((s.get("credits", 0) or 0) for s in subjects)

    subject_results: Dict[str, Dict[str, Any]] = {}
    for subject in subjects:
        subject_id = str(subject["_id"])
        result, code = await _get_subject_results(db, subject_id, institute_id)
        if code == 200:
            subject_results[subject_id] = result

    roster = set()
    for result in subject_results.values():
        roster.update(s["student_id"] for s in result["students"])

    student_result: Dict[str, Dict[str, Any]] = {sid: {"student_id": sid} for sid in roster}

    for subject in subjects:
        subject_id = str(subject["_id"])
        result = subject_results.get(subject_id)
        if not result:
            continue

        subject_name = subject.get("subject_name")
        credits = subject.get("credits", 0) or 0
        by_student = {s["student_id"]: s for s in result["students"]}

        for sid in roster:
            student = by_student.get(sid)
            composite = (student or {}).get("composite_percentage")
            grade = (student or {}).get("course_grade") or "U"
            grade_point = (student or {}).get("grade_point") or 0
            if composite is None:
                composite = 0

            student_result[sid][subject_name] = f"{composite} ({grade})"
            student_result[sid]["overall_total"] = student_result[sid].get("overall_total", 0) + composite
            student_result[sid]["total_credit_points"] = round(
                student_result[sid].get("total_credit_points", 0) + credits * grade_point, 2
            )

    combined_result = list(student_result.values())
    for student in combined_result:
        student["total_credits"] = total_semester_credits

    if not combined_result:
        return {"success": True, "message": "No student results found for this batch and semester", "data": []}, 200

    combined_result.sort(key=lambda x: x["overall_total"], reverse=True)
    for index, student in enumerate(combined_result, start=1):
        student["rank"] = index

    semester_max_marks = 100 * len(subjects) if subjects else 100
    combined_result = assign_relative_grades(combined_result, grading_config, max_marks=semester_max_marks)

    for student in combined_result:
        total_credit_points = round(student.get("total_credit_points", 0), 2)
        total_credits = student.get("total_credits", 0)
        student["total_credit_points"] = total_credit_points
        student["tgpa"] = round(total_credit_points / total_credits, 2) if total_credits else 0
        student["cgpa"] = student["tgpa"]

    return {"success": True, "data": combined_result}, 200


@router.get("/combined-result")
async def combined_result(
    batch_id: str = Query(...), semester: str = Query(...),
    identity: dict = Depends(get_current_identity), db: AsyncIOMotorDatabase = Depends(get_database),
):
    user, institute_id, error = await get_current_user_and_institute(identity, db)
    if error:
        message, code = error
        raise HTTPException(status_code=code, detail=message)

    body, code = await _get_combined_results(db, institute_id, batch_id, semester)
    if code >= 400:
        raise HTTPException(status_code=code, detail=body.get("error"))
    return body


# ============================================================
# COMBINED RESULT EXPORTS (excel + print HTML)
# ============================================================

_RESERVED_COLUMNS = {
    "student_id", "rank", "overall_total", "grade",
    "total_credits", "total_credit_points", "tgpa", "cgpa", "result_source",
}


def _subject_columns_from_rows(rows: List[Dict[str, Any]]) -> List[str]:
    subject_columns: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in _RESERVED_COLUMNS and key not in subject_columns:
                subject_columns.append(key)
    return subject_columns


def _build_combined_result_workbook(rows: List[Dict[str, Any]], semester: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    subject_columns = _subject_columns_from_rows(rows)
    is_transcript = any(row.get("result_source") == "transcript" for row in rows)
    headers = ["Rank", "Student ID"] + subject_columns + ["Overall Total"]
    headers += ["TGPA", "CGPA"] if is_transcript else ["Grade"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Result"

    title = f"Combined Result - Semester {semester}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E78")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_border = Border(
        left=Side(style="thin", color="B7B7B7"), right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"), bottom=Side(style="thin", color="B7B7B7"),
    )

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_index, student in enumerate(rows, start=3):
        values = [student.get("rank"), student.get("student_id")]
        values += [student.get(subject, 0) for subject in subject_columns]
        values += [student.get("overall_total", 0)]
        values += (
            [student.get("tgpa", 0), student.get("cgpa", 0)] if is_transcript else [student.get("grade", "")]
        )

        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    ws.freeze_panes = "A3"

    for col_index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for cell in ws[get_column_letter(col_index)]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_index)].width = min(max_length + 3, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _grade_class(grade: Any) -> str:
    grade = str(grade or "").strip().upper()
    if grade in {"A+", "A", "S/A+"}:
        return "grade-good"
    if grade in {"B+", "B"}:
        return "grade-amber"
    if grade in {"C+", "C"}:
        return "grade-orange"
    return "grade-red"


def _build_combined_result_print_html(rows: List[Dict[str, Any]], batch_id: str, semester: str) -> str:
    subject_columns = _subject_columns_from_rows(rows)
    is_transcript = any(row.get("result_source") == "transcript" for row in rows)
    headers = ["Rank", "Student ID"] + subject_columns + ["Overall Total"]
    headers += ["TGPA", "CGPA"] if is_transcript else ["Grade"]

    def _display(value: Any) -> str:
        return html_module.escape(str(value if value is not None else ""))

    header_html = "".join(f"<th>{html_module.escape(str(h))}</th>" for h in headers)

    row_html = []
    for student in rows:
        rank = student.get("rank")
        rank_class = {1: "rank-one", 2: "rank-two", 3: "rank-three"}.get(rank, "")

        subject_cells = "".join(
            f'<td class="numeric">{_display(student.get(subject, 0))}</td>' for subject in subject_columns
        )
        grade = student.get("grade", "")
        if is_transcript:
            result_cells = (
                f'<td class="overall-total">{_display(student.get("overall_total", 0))}</td>'
                f'<td class="numeric">{_display(student.get("tgpa", 0))}</td>'
                f'<td class="numeric">{_display(student.get("cgpa", 0))}</td>'
            )
        else:
            result_cells = (
                f'<td class="overall-total">{_display(student.get("overall_total", 0))}</td>'
                f'<td><span class="grade-badge {_grade_class(grade)}">{_display(grade)}</span></td>'
            )

        row_html.append(f"""
            <tr class="{rank_class}">
                <td class="rank-cell">{_display(rank)}</td>
                <td class="student-id">{_display(student.get("student_id"))}</td>
                {subject_cells}
                {result_cells}
            </tr>""")

    table_or_empty = (
        f'<table><thead><tr>{header_html}</tr></thead><tbody>{"".join(row_html)}</tbody></table>'
        if rows else '<div class="empty-state">No Result Found</div>'
    )

    generated_at = datetime.now(timezone.utc).strftime("%d-%m-%Y %I:%M %p")

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Combined Result - Semester {html_module.escape(str(semester))}</title>
<style>
  @page {{ size: A4 portrait; margin: 14mm 12mm 18mm 12mm; }}
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; background: #eef1f5; color: #1f2937; font-family: "Segoe UI", Arial, sans-serif; line-height: 1.4; font-size: 12px; }}
  .page {{ width: min(100%, 900px); margin: 24px auto; padding: 26px 30px 20px; background: #fff; box-shadow: 0 18px 45px rgba(15,23,42,.12); }}
  h1 {{ font-size: 18px; margin: 0 0 4px; }}
  .meta {{ color: #6b7280; font-size: 11px; margin-bottom: 8px; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
  th, td {{ border: 1px solid #d0d5dd; padding: 6px 8px; text-align: center; font-size: 11px; }}
  th {{ background: #d9eaf7; }}
  td.student-id {{ text-align: left; font-weight: 600; }}
  td.overall-total {{ font-weight: 700; }}
  tr.rank-one td.rank-cell {{ color: #b8860b; font-weight: 700; }}
  tr.rank-two td.rank-cell {{ color: #708090; font-weight: 700; }}
  tr.rank-three td.rank-cell {{ color: #8b5a2b; font-weight: 700; }}
  .grade-badge {{ padding: 2px 8px; border-radius: 10px; font-weight: 600; }}
  .grade-good {{ background: #dcfce7; color: #166534; }}
  .grade-amber {{ background: #fef9c3; color: #854d0e; }}
  .grade-orange {{ background: #ffedd5; color: #9a3412; }}
  .grade-red {{ background: #fee2e2; color: #991b1b; }}
  .empty-state {{ text-align: center; padding: 40px; color: #6b7280; }}
  @media print {{ body {{ background: #fff; }} .page {{ box-shadow: none; margin: 0; }} }}
</style>
</head>
<body>
  <div class="page">
    <h1>Combined Result - Semester {html_module.escape(str(semester))}</h1>
    <div class="meta">Batch: {html_module.escape(str(batch_id))} &middot; Generated {generated_at}</div>
    {table_or_empty}
  </div>
</body>
</html>"""


@router.get("/combined-result/export")
async def combined_result_export(
    batch_id: str = Query(...), semester: str = Query(...),
    identity: dict = Depends(get_current_identity), db: AsyncIOMotorDatabase = Depends(get_database),
):
    user, institute_id, error = await get_current_user_and_institute(identity, db)
    if error:
        message, code = error
        raise HTTPException(status_code=code, detail=message)

    body, code = await _get_combined_results(db, institute_id, batch_id, semester)
    if code != 200:
        raise HTTPException(status_code=code, detail=body.get("error"))

    rows = body.get("data", [])
    if not rows:
        raise HTTPException(status_code=404, detail="No combined result data found")

    xlsx_bytes = await asyncio.to_thread(_build_combined_result_workbook, rows, semester)

    filename = f"combined_result_semester_{semester}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/combined-result/print")
async def combined_result_print(
    batch_id: str = Query(...), semester: str = Query(...),
    identity: dict = Depends(get_current_identity), db: AsyncIOMotorDatabase = Depends(get_database),
):
    user, institute_id, error = await get_current_user_and_institute(identity, db)
    if error:
        message, code = error
        raise HTTPException(status_code=code, detail=message)

    body, code = await _get_combined_results(db, institute_id, batch_id, semester)
    if code != 200:
        raise HTTPException(status_code=code, detail=body.get("error"))

    rows = body.get("data", [])
    html_content = _build_combined_result_print_html(rows, batch_id, semester)

    return Response(
        content=html_content, media_type="text/html",
        headers={"Content-Disposition": f"inline; filename=combined_result_semester_{semester}.html"},
    )


# ============================================================
# SINGLE-EXAM DETAILED EXCEL REPORT
# ============================================================

def _col_idx_to_letter(col_idx: int) -> str:
    if col_idx < 26:
        return chr(65 + col_idx)
    return chr(64 + col_idx // 26) + chr(65 + col_idx % 26)


def _safe_float(val: Any) -> float:
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _build_detailed_excel_workbook(folder: Dict[str, Any], eval_details: List[Dict[str, Any]], answers: List[Dict[str, Any]]) -> bytes:
    import xlsxwriter

    question_paper = folder.get("question_paper", {}) or {}
    total_questions = int(question_paper.get("no_of_questions", 0) or 0)

    first_qw: Dict[Any, Dict[str, Any]] = {}
    for q in answers[0].get("questionwise_marking", []):
        first_qw[q.get("question_no")] = q

    question_layout = []
    for q_idx in range(total_questions):
        q_no = q_idx + 1
        q_data = first_qw.get(q_no, {})

        if q_data and q_data.get("max_marks") not in (None, ""):
            q_max = _safe_float(q_data.get("max_marks"))
        elif q_idx < len(eval_details):
            q_max = _safe_float(eval_details[q_idx].get("maxMarks", 0))
        else:
            q_max = 0

        cos = []
        for c in q_data.get("cos", []):
            co_code = c.get("co_code", "")
            co_max = _safe_float(c.get("max_marks", 0))
            if co_code:
                cos.append({"co_code": co_code, "max_marks": co_max})

        question_layout.append({"q_no": q_no, "max_marks": q_max, "cos": cos})

    question_col_count = sum(1 + len(q["cos"]) for q in question_layout)
    summary_col_count = 5
    total_cols = 3 + question_col_count + summary_col_count
    summary_start_col = 3 + question_col_count

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet("Evaluation Report")

    title_format = workbook.add_format({
        "bold": True, "font_size": 16, "font_color": "#333333",
        "align": "center", "valign": "vcenter", "border": 1,
    })
    max_marks_format = workbook.add_format({
        "bold": True, "font_size": 11, "font_color": "#FA7D00",
        "bg_color": "#F2F2F2", "align": "center", "border": 1,
    })
    header_gray_format = workbook.add_format({
        "bold": True, "font_size": 11, "font_color": "#FFFFFF",
        "bg_color": "#4A4A4A", "align": "center", "valign": "vcenter", "border": 1,
    })
    header_ems_format = workbook.add_format({
        "bold": True, "font_size": 11, "font_color": "#FFFFFF",
        "bg_color": "#E26B0A", "align": "center", "valign": "vcenter", "border": 1,
    })
    header_faculty_format = workbook.add_format({
        "bold": True, "font_size": 11, "font_color": "#FFFFFF",
        "bg_color": "#76933C", "align": "center", "valign": "vcenter", "border": 1,
    })
    cell_center_format = workbook.add_format({"align": "center", "valign": "vcenter", "border": 1, "border_color": "#BFBFBF"})
    cell_left_format = workbook.add_format({"align": "left", "valign": "vcenter", "border": 1, "border_color": "#BFBFBF"})
    date_format = workbook.add_format({
        "align": "center", "valign": "vcenter", "border": 1, "border_color": "#BFBFBF", "num_format": "dd-mm-yyyy",
    })

    worksheet.set_column("A:A", 5.66)
    worksheet.set_column("B:B", 17.0)
    worksheet.set_column("C:C", 20.0)

    col = 3
    for q_info in question_layout:
        worksheet.set_column(f"{_col_idx_to_letter(col)}:{_col_idx_to_letter(col)}", 13.0)
        col += 1
        for _ in q_info["cos"]:
            worksheet.set_column(f"{_col_idx_to_letter(col)}:{_col_idx_to_letter(col)}", 13.0)
            col += 1

    for offset in range(summary_col_count):
        ltr = _col_idx_to_letter(summary_start_col + offset)
        width = 15.0 if offset == 4 else 13.0
        worksheet.set_column(f"{ltr}:{ltr}", width)

    last_col_letter = _col_idx_to_letter(total_cols - 1)
    worksheet.merge_range(
        f"A1:{last_col_letter}1", f"Evaluation Report Analysis - {folder.get('foldername', 'Untitled')}", title_format
    )
    worksheet.set_row(0, 28.05)

    worksheet.merge_range("A2:C2", "Max marks per question->", max_marks_format)

    col = 3
    for q_info in question_layout:
        worksheet.write(f"{_col_idx_to_letter(col)}2", f"Max marks -{q_info['max_marks']}", max_marks_format)
        col += 1
        for co in q_info["cos"]:
            worksheet.write(f"{_col_idx_to_letter(col)}2", f"CO Max -{co['max_marks']}", max_marks_format)
            col += 1

    for offset in range(summary_col_count):
        worksheet.write(f"{_col_idx_to_letter(summary_start_col + offset)}2", "", max_marks_format)

    worksheet.write("A3", "Sr. No", header_gray_format)
    worksheet.write("B3", "File Name", header_gray_format)
    worksheet.write("C3", "Student Name", header_gray_format)

    col = 3
    for q_info in question_layout:
        q_no = q_info["q_no"]
        worksheet.write(f"{_col_idx_to_letter(col)}3", f"EMS Ques-{q_no}", header_ems_format)
        col += 1
        for co in q_info["cos"]:
            worksheet.write(f"{_col_idx_to_letter(col)}3", co["co_code"], header_faculty_format)
            col += 1

    summary_headers = ["EMS Total", "Faculty Total", "Max Marks", "Percentage", "Evaluated At"]
    summary_formats = [header_ems_format, header_faculty_format, header_gray_format, header_gray_format, header_gray_format]
    for offset, (header, fmt) in enumerate(zip(summary_headers, summary_formats)):
        worksheet.write(f"{_col_idx_to_letter(summary_start_col + offset)}3", header, fmt)

    worksheet.set_row(2, 22.05)
    worksheet.freeze_panes(3, 3)

    row = 3
    for idx, answer in enumerate(answers, start=1):
        worksheet.write(row, 0, idx, cell_center_format)
        worksheet.write(row, 1, answer.get("filename", "Unknown"), cell_left_format)
        worksheet.write(row, 2, answer.get("student_name", "-"), cell_left_format)

        qw_map: Dict[Any, Dict[str, Any]] = {}
        for q in answer.get("questionwise_marking", []):
            qw_map[q.get("question_no")] = q

        col = 3
        for q_info in question_layout:
            q_no = q_info["q_no"]
            q_data = qw_map.get(q_no, {})
            flags = q_data.get("flags", {}) if q_data else {}
            is_unanswered = flags.get("unanswered", False) or not q_data

            if is_unanswered:
                worksheet.write(row, col, "Not Attempted", cell_center_format)
            else:
                final_marks = q_data.get("final_marks")
                if final_marks is None or final_marks == "":
                    worksheet.write(row, col, "Not Attempted", cell_center_format)
                else:
                    worksheet.write(row, col, final_marks, cell_center_format)
            col += 1

            answer_co_map: Dict[Any, Dict[str, Any]] = {}
            if q_data:
                for c in q_data.get("cos", []):
                    answer_co_map[c.get("co_code")] = c

            for co in q_info["cos"]:
                if is_unanswered:
                    worksheet.write(row, col, "Not Attempted", cell_center_format)
                else:
                    c_data = answer_co_map.get(co["co_code"])
                    if c_data is None:
                        worksheet.write(row, col, "-", cell_center_format)
                    else:
                        val = c_data.get("final_co_marks")
                        if val is None or val == "":
                            val = c_data.get("ai_marks", 0)
                        try:
                            worksheet.write(row, col, float(val), cell_center_format)
                        except (ValueError, TypeError):
                            worksheet.write(row, col, "-", cell_center_format)
                col += 1

        total_ai = answer.get("total_ai_marks", 0) or 0
        total_final = answer.get("total_final_marks", total_ai) or 0
        max_marks = answer.get("total_max_marks", 0) or 0

        worksheet.write(row, summary_start_col, total_ai, cell_center_format)
        worksheet.write(row, summary_start_col + 1, total_final, cell_center_format)
        worksheet.write(row, summary_start_col + 2, max_marks, cell_center_format)

        if max_marks > 0:
            worksheet.write(row, summary_start_col + 3, f"{(total_final / max_marks * 100):.2f}%", cell_center_format)
        else:
            worksheet.write(row, summary_start_col + 3, "-", cell_center_format)

        evaluated_at = answer.get("evaluated_at")
        written = False
        if evaluated_at:
            if isinstance(evaluated_at, str):
                try:
                    dt = datetime.fromisoformat(evaluated_at.replace("Z", "+00:00"))
                    dt += timedelta(hours=5, minutes=30)
                    worksheet.write_datetime(
                        row, summary_start_col + 4, dt.replace(hour=0, minute=0, second=0, microsecond=0), date_format
                    )
                    written = True
                except Exception:
                    pass
            elif isinstance(evaluated_at, datetime):
                dt = evaluated_at + timedelta(hours=5, minutes=30)
                worksheet.write_datetime(
                    row, summary_start_col + 4, dt.replace(hour=0, minute=0, second=0, microsecond=0), date_format
                )
                written = True
        if not written:
            worksheet.write(row, summary_start_col + 4, "-", cell_center_format)

        row += 1

    workbook.close()
    output.seek(0)
    return output.getvalue()


@router.get("/download-detailed-excel/{folder_id}")
async def download_detailed_excel(
    folder_id: str,
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    if not ObjectId.is_valid(folder_id):
        raise HTTPException(status_code=400, detail="Invalid folder id")

    user, faculty_id, error = await get_current_user_and_faculty_details(identity, db)
    if error:
        message, code = error
        raise HTTPException(status_code=code, detail=message)

    exam_object_id = ObjectId(folder_id)
    folder = await db["newsavedDocs"].find_one({"_id": exam_object_id, "faculty_id": ObjectId(faculty_id)})
    if not folder:
        raise HTTPException(status_code=404, detail="Folder not found or unauthorized")

    evaluation_doc = await db["evaluationDetails"].find_one({"exam_id": exam_object_id})
    eval_details = (evaluation_doc or {}).get("evaluation_details", [])

    answers = [a async for a in db["answerDetails"].find({"exam_id": exam_object_id})]
    if not answers:
        raise HTTPException(status_code=404, detail="No answer scripts found")

    xlsx_bytes = await asyncio.to_thread(_build_detailed_excel_workbook, folder, eval_details, answers)

    filename = f"{folder.get('folder_name', 'Evaluation_Report_Analysis')}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ============================================================
# CO-PO ATTAINMENT EXCEL REPORT
# Ported from controllers/institute/coExcel_controller.py — see
# app/services/co_excel.py for the workbook-building logic (kept in its
# own module; it's ~900 lines of pure openpyxl construction).
#
# Flask original had no institute/subject ownership check on this
# endpoint — subject_id was trusted as-is from the URL. Fixed here:
# scoped to the caller's institute (resolved via resolve_current_institute_id,
# which covers both institute admins and faculty of that institute), matching
# the ownership pattern already used elsewhere in this router/institute_hierarchy.py.
# ============================================================

@router.get("/co-detailed-excel/{subject_id}")
async def download_co_detailed_excel(
    subject_id: str,
    identity: dict = Depends(get_current_identity),
    db: AsyncIOMotorDatabase = Depends(get_database),
):
    if not ObjectId.is_valid(subject_id):
        raise HTTPException(status_code=400, detail="Invalid subject_id")

    institute_id = await resolve_current_institute_id(identity, db)
    subject, err, code = await validate_entity_ownership(db["subjectDetails"], subject_id, institute_id)
    if err:
        raise HTTPException(status_code=code, detail=err["error"])

    try:
        data = await fetch_co_report_data(db, subject_id)
    except CoReportError as e:
        raise HTTPException(status_code=e.status_code, detail=e.payload)

    xlsx_bytes = await asyncio.to_thread(build_co_report_workbook, data)

    filename = f"CO_Report_{subject_id}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
