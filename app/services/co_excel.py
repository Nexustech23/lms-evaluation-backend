# ============================================================
# CO-PO ATTAINMENT EXCEL REPORT
#
# Layout mirrors the institute's standard NBA-style CO-PO attainment
# workbook (Start / one sheet per exam / Final CO Attainment / Final PO
# attainment), populated with this subject's real exam, marks and CO-PO
# data. Sheet structure, colors (extracted from that workbook's theme.xml)
# and fonts (Times New Roman 12pt) are matched deliberately so faculty
# get a familiar, audit-ready report instead of an ad-hoc format.
#
# fetch_exams/fetch_targets/fetch_co_po/fetch_subject_meta (+ _try_find)
# touch the DB (async Motor). Sheet builders are pure/sync openpyxl —
# build_co_report_workbook() runs via asyncio.to_thread from the router.
#
# NOTE: no institute/subject ownership check here — subject_id is trusted
# as-is from the URL, same as the rest of this module historically.
# ============================================================

import io
from typing import Any, Dict, List, Optional, Tuple

from bson import ObjectId
from motor.motor_asyncio import AsyncIOMotorDatabase
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class CoReportError(Exception):
    """Raised for guard conditions the router turns into (400/404, payload) responses."""

    def __init__(self, status_code: int, payload: Dict[str, Any]):
        self.status_code = status_code
        self.payload = payload
        super().__init__(payload.get("error", "CO report error"))


# ═══════════════════════════════════════════════════════════════════════
#  PALETTE — extracted from the reference workbook's theme.xml
#  (accent1=4472C4, accent2=ED7D31, accent5=5B9BD5, dk2=44546A) with
#  Excel's tint formula applied, plus its solid (non-theme) fills.
# ═══════════════════════════════════════════════════════════════════════
FONT_NAME = "Times New Roman"
FONT_SIZE = 12


def _font(bold=False, color="000000", size=FONT_SIZE, italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, italic=italic, color=color)


BOLD = _font(bold=True)
WHITE_BOLD = _font(bold=True, color="FFFFFF")

LABEL_ORANGE = PatternFill("solid", fgColor="FBE4D5")    # Start label cells
LABEL_BLUE = PatternFill("solid", fgColor="DEEAF6")       # Target label / CO-PO matrix data cells
BRIGHT_BLUE = PatternFill("solid", fgColor="00B0F0")      # CO-PO header row / "Target" bar
YELLOW = PatternFill("solid", fgColor="FFFF00")           # criteria label row / Average row / TOTAL header
MED_ORANGE = PatternFill("solid", fgColor="F4B083")       # criteria comparison-mark value cells (Start)
DARK_BLUE = PatternFill("solid", fgColor="4472C4")        # Q.No/COs/Marks label column
LIGHT_BLUE_DATA = PatternFill("solid", fgColor="B4C6E7")  # per-question data columns / CO Wise Marks block
ROSTER_BLUE = PatternFill("solid", fgColor="8DB4E2")      # Sl.No/Name/Reg.No header
DARK_NAVY = PatternFill("solid", fgColor="44546A")        # "CO Wise Percentage" header
GREEN_HDR = PatternFill("solid", fgColor="92D050")        # criteria-block header + value cells (per-exam sheets)
GREY = PatternFill("solid", fgColor="DADADA")             # final Attained/Not-Attained verdict cells

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = BORDER


def short_exam_label(folder_name: str, course_name: str) -> str:
    """'Digital Logic Design-CA1' + course 'Digital Logic Design' -> 'CA1'.
    Falls back to the full folder_name if it doesn't follow the
    '<course><sep><type>' convention, since not every subject will."""
    name = (folder_name or "").strip()
    course = (course_name or "").strip()
    if course and name.lower().startswith(course.lower()):
        rest = name[len(course):].lstrip(" -–—:_")
        if rest:
            return rest
    return name or "Exam"


# Groups multiple sub-instruments (e.g. CT-1/CT-2/CT-3, Quiz-1..5) into one
# reference-style category column in the Final CO Attainment contribution
# table, matching "Class Test / MCQ/Quiz / Laboratory/Assignment/Activity /
# MSA / ESA / CES" from the reference workbook. Any exam_type that doesn't
# match a known prefix falls back to being its own category (1:1) — this
# keeps the simpler single-exam-per-category case (e.g. a subject that
# just has CA1/CA2/CA3/Mid-Term/End-Term, no sub-instruments) working
# exactly as before, unchanged.
_CATEGORY_PREFIXES = [
    ("CT", "Class Test"),
    ("QUIZ", "MCQ/Quiz"),
    ("ASS", "Laboratory/Assignment/Activity"),
    ("MOOCS", "MOOCs"),
    ("MSA", "Mid Semester Assessment"),
    ("ESA", "End Semester Assessment"),
    ("CES", "Course Exit Survey"),
]

# Reference-workbook sheet names for each category (Class Tests/Quiz/Assignment/
# MOOCs/Mid-Sem/End-Sem/CES hold ALL of that category's sub-instruments side by
# side in one sheet, exactly like the source workbook — not one sheet per exam).
CATEGORY_SHEET_NAMES = {
    "Class Test": "Class Tests",
    "MCQ/Quiz": "Quiz",
    "Laboratory/Assignment/Activity": "Assignment",
    "MOOCs": "MOOCs",
    "Mid Semester Assessment": "Mid-Sem",
    "End Semester Assessment": "End-Sem",
    "Course Exit Survey": "CES",
}


def exam_category(exam_report: Dict[str, Any]) -> str:
    exam_type = (exam_report.get("exam_type") or "").upper()
    for prefix, label in _CATEGORY_PREFIXES:
        if exam_type.startswith(prefix):
            return label
    return exam_report.get("folder_name") or exam_report.get("exam_id", "Exam")


def autofit_columns(ws, min_width=8, max_width=32, padding=2):
    """Sets each column's width from its longest cell value, so long labels
    (exam names, attainment-rule text, etc.) don't get silently truncated —
    generalizes across whatever real subject/exam names a report contains."""
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            widths[cell.column] = max(widths.get(cell.column, 0), length)
    for col_idx, length in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + padding, min_width), max_width)


def level_for_pct(pct, bands):
    """Matches the reference workbook's cascading IF formula exactly:
    IF(pct>=70,3, IF(pct>=60,2, IF(pct>=40,1, 1))) — the floor for any
    valid percentage is the LOWEST configured band's level, never 0.
    A 0/None result only ever comes from the caller's own "no data at
    all" guard (has_data), matching the formula's IFERROR branch."""
    if not bands:
        return 0
    for band in bands:
        if band["min_percentage"] <= pct <= band["max_percentage"]:
            return band["level"]
    return min(bands, key=lambda b: b["min_percentage"])["level"]


def _level_if_formula(pct_ref: str, targets: List[Dict[str, Any]]) -> str:
    """Builds the Excel-formula equivalent of level_for_pct() as a nested IF
    referencing a percentage cell, so the workbook shows a real formula
    (not a static number) when a faculty member clicks the Level cell."""
    if not targets:
        return "0"
    # Build from the lowest band outward so the HIGHEST threshold ends up as
    # the outermost (first-checked) IF — e.g. IF(pct>=70,3,IF(pct>=60,2,
    # IF(pct>=40,1,1))), matching the reference workbook's cascade exactly.
    sorted_bands_asc = sorted(targets, key=lambda b: b["min_percentage"])
    floor_level = sorted_bands_asc[0]["level"]
    expr = str(floor_level)
    for band in sorted_bands_asc:
        expr = f'IF({pct_ref}>={band["min_percentage"]},{band["level"]},{expr})'
    return expr


def criteria_labels():
    return [
        (">=70% of students scoring>=", 3),
        ("60%-70% of students scoring>=", 2),
        ("40%-60%of students scoring>=", 1),
    ]


# ═══════════════════════════════════════════════════════════════════════
#  BUSINESS LOGIC (pure — no DB dependency)
# ═══════════════════════════════════════════════════════════════════════
def calc_pct(obtained, max_marks):
    if not max_marks:
        return None
    return round((obtained / max_marks) * 100, 2)


def aggregate_by_co(student):
    agg = {}
    for q in student.get("questions", []):
        for co in q.get("cos", []):
            code = co["co_code"]
            if code not in agg:
                agg[code] = {"obtained": 0, "max": 0}
            agg[code]["obtained"] += co.get("obtained_marks", 0) or 0
            agg[code]["max"] += co.get("max_marks", 0) or 0
    return agg


def build_columns(students):
    """Sort questions numerically, not as strings (1,2,10 not 1,10,2)."""
    col_map = {}
    for student in students:
        for q in student.get("questions", []):
            for co in q.get("cos", []):
                key = f"{q['question_no']}|{co['co_code']}"
                if key not in col_map:
                    col_map[key] = {
                        "questionNo": q["question_no"],
                        "coCode": co["co_code"],
                        "maxMarks": co.get("max_marks", 0),
                    }

    def sort_key(c):
        try:
            return (int(c["questionNo"]), c["coCode"])
        except Exception:
            return (9999, str(c["questionNo"]) + c["coCode"])

    return sorted(col_map.values(), key=sort_key)


def get_co_level_info(student_pcts, targets):
    if not student_pcts or not targets:
        return None
    comparison_pct = targets[0]["comparision_percentage"]
    valid = [p for p in student_pcts if p is not None]
    total = len(valid)
    students_above = sum(1 for p in valid if p >= comparison_pct)
    achievement_pct = round((students_above / total) * 100, 2) if total else 0
    matched = next(
        (t for t in targets if t["min_percentage"] <= achievement_pct <= t["max_percentage"]), None
    )
    return {
        "achievementPct": achievement_pct, "level": matched["level"] if matched else 0,
        "comparisionPct": comparison_pct, "studentsAbove": students_above, "total": total,
    }


def build_direct_attainment(exams, targets, co_attainment_target=2):
    """Per-CO, per-exam level + weighted contribution, and the final weighted sum."""
    co_set = set()
    for exam in exams:
        for student in exam.get("students", []):
            for q in student.get("questions", []):
                for co in q.get("cos", []):
                    co_set.add(co["co_code"])
    all_cos = sorted(co_set)

    level_matrix: Dict[str, Any] = {}
    final_att: Dict[str, float] = {}
    for co in all_cos:
        level_matrix[co] = {}
        final_att[co] = 0.0
        for exam in exams:
            student_pcts = []
            for student in exam.get("students", []):
                agg = aggregate_by_co(student)
                d = agg.get(co)
                student_pcts.append(calc_pct(d["obtained"], d["max"]) if d else None)
            info = get_co_level_info(student_pcts, targets)
            level = info["level"] if info else 0
            has_data = bool(info and info["total"] > 0)
            weightage = float(exam.get("weightage", 0) or 0)
            weighted = level * (weightage / 100) if has_data else 0.0
            level_matrix[co][exam["exam_id"]] = {
                "level": level, "weighted": weighted, "hasData": has_data,
                "achievementPct": info["achievementPct"] if info else None,
            }
            final_att[co] += weighted
        final_att[co] = round(final_att[co] * 100) / 100

    return {"allCOs": all_cos, "levelMatrix": level_matrix, "finalAttainment": final_att}


# ═══════════════════════════════════════════════════════════════════════
#  DATA FETCHING (async — Motor)
# ═══════════════════════════════════════════════════════════════════════
async def _try_find(db: AsyncIOMotorDatabase, collection_name: str, field: str, eval_id):
    docs = [d async for d in db[collection_name].find({field: eval_id})]
    if not docs:
        alt = str(eval_id) if isinstance(eval_id, ObjectId) else ObjectId(str(eval_id))
        docs = [d async for d in db[collection_name].find({field: alt})]
    return docs


def _safe_float(v, default=0.0):
    try:
        return float(v)
    except Exception:
        return default


def _normalise_answer(ans):
    name = (ans.get("student_name") or ans.get("studentName") or ans.get("name") or "").strip()
    fname = (ans.get("filename") or ans.get("file_name") or ans.get("originalName") or "").strip()
    raw_qs = (
        ans.get("questionwise_marking") or ans.get("questions") or ans.get("questionDetails")
        or ans.get("question_details") or []
    )
    questions = []
    for q in raw_qs:
        q_no = q.get("question_no") or q.get("questionNo") or q.get("q_no") or q.get("number") or 0
        raw_cos = (
            q.get("co_marks") or q.get("cos") or q.get("co_details") or q.get("coDetails")
            or q.get("course_outcomes") or []
        )
        cos = []
        for co in raw_cos:
            code = co.get("co_code") or co.get("coCode") or co.get("code") or ""
            obtained = _safe_float(
                co.get("final_co_marks") if co.get("final_co_marks") is not None
                else co.get("ai_marks") if co.get("ai_marks") is not None
                else co.get("obtained_marks") if co.get("obtained_marks") is not None
                else co.get("obtainedMarks") if co.get("obtainedMarks") is not None
                else co.get("final_marks") if co.get("final_marks") is not None
                else co.get("marks")
            )
            max_m = _safe_float(
                co.get("max_marks") if co.get("max_marks") is not None
                else co.get("maxMarks") if co.get("maxMarks") is not None
                else co.get("maximum_marks")
            )
            if code:
                cos.append({"co_code": code, "obtained_marks": obtained, "max_marks": max_m})
        if cos:
            questions.append({"question_no": q_no, "cos": cos})
    return {"answer_id": str(ans["_id"]), "student_name": name, "filename": fname, "questions": questions}


async def fetch_exams(db: AsyncIOMotorDatabase, subject_id: str) -> List[Dict[str, Any]]:
    evaluations = [e async for e in db["newsavedDocs"].find({"subject_id": ObjectId(subject_id)})]
    exams = []
    for ev in evaluations:
        raw = await _try_find(db, "answerDetails", "exam_id", ev["_id"])
        if not raw:
            raw = await _try_find(db, "answerDetails", "evaluation_id", ev["_id"])
        if not raw:
            raw = await _try_find(db, "answerDetails", "newsaved_id", ev["_id"])
        students = [_normalise_answer(a) for a in raw]
        students.sort(key=lambda s: (s["student_name"].lower() if s["student_name"] else s["filename"].lower()))
        exams.append({
            "exam_id": str(ev["_id"]),
            "folder_name": ev.get("folder_name", ""),
            "weightage": _safe_float(ev.get("weightage", 0)),
            "is_course_exit_summary": bool(ev.get("is_course_exit_summary", False)),
            "exam_type": ev.get("exam_type", ""),
            "students": students,
        })
    # weightage ascending (CAs before Mid/End-Term), folder_name as tiebreak
    # among equal-weight exams — mirrors the natural CA1→CA2→CA3→Mid→End order.
    exams.sort(key=lambda e: (e["weightage"], (e.get("folder_name") or "").lower()))
    return exams


async def fetch_targets(db: AsyncIOMotorDatabase, subject_id: str) -> Tuple[List[Dict[str, Any]], float]:
    subject = await db["subjectDetails"].find_one({"_id": ObjectId(subject_id)})
    if not subject:
        return [], 1.8
    prog_id = subject.get("programme_id")
    prog = None
    if prog_id:
        prog = await db["programmeDetails"].find_one({"_id": prog_id})
    targets = []
    if prog:
        targets = prog.get("targets", []) or prog.get("po_targets", []) or []
    co_attainment_target = (
        _safe_float(prog.get("coAttainmentTarget"), 1.8)
        if prog and prog.get("coAttainmentTarget") is not None
        else _safe_float(subject.get("co_attainment_target", subject.get("coAttainmentTarget", 1.8)), 1.8)
    )
    return targets, co_attainment_target


async def fetch_co_po(db: AsyncIOMotorDatabase, subject_id: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    subject = await db["subjectDetails"].find_one({"_id": ObjectId(subject_id)})
    if not subject:
        return {}, []
    return subject.get("co_po_matrix", {}) or {}, subject.get("co_po_matrix_arr", []) or []


async def fetch_subject_meta(db: AsyncIOMotorDatabase, subject_id: str) -> Dict[str, Any]:
    """Real header info for the Start sheet — course/faculty/branch/semester.
    Missing fields are left blank rather than guessed."""
    subject = await db["subjectDetails"].find_one({"_id": ObjectId(subject_id)}) or {}
    meta = {
        "course": subject.get("subject_name", ""),
        "course_code": subject.get("subject_code") or "N/A",
        "semester": subject.get("semester", ""),
        "co_descriptions": {c["co_code"]: c.get("description", "") for c in (subject.get("co") or [])},
        "branch": "", "faculty_display": "",
    }
    prog_id = subject.get("programme_id")
    if prog_id:
        prog = await db["programmeDetails"].find_one({"_id": prog_id})
        if prog:
            meta["branch"] = prog.get("programme_name", "")
    faculty_id = subject.get("faculty_id")
    if faculty_id:
        fac = await db["facultyDetails"].find_one({"_id": faculty_id})
        if fac:
            designation = fac.get("designation", "")
            user = await db["users"].find_one({"_id": fac.get("user_id")}) if fac.get("user_id") else None
            email = (user or {}).get("email", "")
            name = (user or {}).get("name", "")
            meta["faculty_display"] = " — ".join(p for p in (name or email, designation) if p)
    return meta


def _po_sort_key(code: str):
    """Numeric-aware so PO2 sorts before PO10, and PSOs follow POs."""
    import re
    m = re.match(r"(PSO|PO)(\d+)", code)
    if not m:
        return (2, 0, code)
    prefix, num = m.groups()
    return (1 if prefix == "PSO" else 0, int(num), code)


def po_codes_from_matrix(co_po_matrix_obj: Dict[str, Any]) -> List[str]:
    po_set = set()
    for po_map in (co_po_matrix_obj or {}).values():
        for po in (po_map or {}).keys():
            po_set.add(po)
    return sorted(po_set, key=_po_sort_key)


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: Start
# ═══════════════════════════════════════════════════════════════════════
def build_start_sheet(wb, meta, targets, co_attainment_target, co_po_matrix_obj, all_cos):
    ws = wb.create_sheet("Start")

    ws["A6"] = "Name:"; ws.merge_cells("A6:B6")
    ws["C6"] = meta.get("faculty_display", ""); ws.merge_cells("C6:F6")
    ws["H6"] = "Target for CO attainment:"; ws.merge_cells("H6:J6"); ws["H6"].fill = LABEL_BLUE
    comparison_pct = targets[0]["comparision_percentage"] if targets else None
    ws["K6"] = comparison_pct
    ws["L6"] = co_attainment_target

    ws["A7"] = "Designation:"; ws.merge_cells("A7:B7")
    ws["C7"] = ""; ws.merge_cells("C7:F7")
    ws["A8"] = "Course:"; ws.merge_cells("A8:B8")
    ws["C8"] = meta.get("course", ""); ws.merge_cells("C8:F8")
    ws["H8"] = "Target"; ws.merge_cells("H8:Q8"); ws["H8"].fill = BRIGHT_BLUE; ws["H8"].font = WHITE_BOLD
    ws["A9"] = "Course Code:"; ws.merge_cells("A9:B9")
    ws["C9"] = meta.get("course_code", ""); ws.merge_cells("C9:F9")

    for i, (label, level) in enumerate(criteria_labels()):
        r = 9 + i
        ws.cell(row=r, column=8, value=label)
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)
        ws.cell(row=r, column=12, value=comparison_pct)
        ws.cell(row=r, column=13, value="marks")
        ws.cell(row=r, column=15, value=level)
        ws.merge_cells(start_row=r, start_column=15, end_row=r, end_column=17)
        ws.cell(row=r, column=8).fill = YELLOW
        ws.cell(row=r, column=12).fill = MED_ORANGE
        ws.cell(row=r, column=15).fill = LABEL_BLUE

    ws["A10"] = "Branch/section:"; ws.merge_cells("A10:B10")
    ws["C10"] = meta.get("branch", ""); ws.merge_cells("C10:F10")
    ws["A11"] = "Semester:"; ws.merge_cells("A11:B11")
    ws["C11"] = str(meta.get("semester", "")); ws.merge_cells("C11:F11")
    ws["A12"] = "Academic Year:"; ws.merge_cells("A12:B12")
    ws["C12"] = ""; ws.merge_cells("C12:F12")

    for coord in ("A6", "A7", "A8", "A9", "A10", "A11", "A12", "C6", "C7", "C8", "C9", "C10", "C11", "C12"):
        ws[coord].fill = LABEL_ORANGE
        ws[coord].font = BOLD
    border_range(ws, 6, 12, 1, 17)

    ws["B14"] = "CO-PO MAPPING" + ("" if co_po_matrix_obj else "  (no CO-PO matrix has been entered for this subject yet)")
    ws["B14"].font = _font(bold=True, italic=not bool(co_po_matrix_obj), color="B45309" if not co_po_matrix_obj else "000000")
    ws["B14"].fill = LABEL_BLUE

    pos = po_codes_from_matrix(co_po_matrix_obj)
    if pos and all_cos:
        ws["B16"] = "CO's"
        for i, po in enumerate(pos):
            ws.cell(row=16, column=3 + i, value=po)
        for c in range(2, 3 + len(pos)):
            ws.cell(row=16, column=c).fill = BRIGHT_BLUE
            ws.cell(row=16, column=c).font = WHITE_BOLD

        for r, co in enumerate(all_cos):
            row = 17 + r
            ws.cell(row=row, column=2, value=co).font = BOLD
            for i, po in enumerate(pos):
                val = (co_po_matrix_obj.get(co) or {}).get(po)
                cell = ws.cell(row=row, column=3 + i, value=val if val is not None else "")
                cell.fill = LABEL_BLUE
                cell.alignment = CENTER

        avg_row = 17 + len(all_cos)
        ws.cell(row=avg_row, column=2, value="Average").font = BOLD
        ws.cell(row=avg_row, column=2).fill = YELLOW
        for i, po in enumerate(pos):
            col_letter = get_column_letter(3 + i)
            cell = ws.cell(
                row=avg_row, column=3 + i,
                value=f"=IFERROR(AVERAGE({col_letter}17:{col_letter}{avg_row - 1}),\"\")",
            )
            cell.font = BOLD
            cell.fill = YELLOW
            cell.alignment = CENTER
        border_range(ws, 16, avg_row, 2, 2 + len(pos))

    for col, w in {"A": 16, "B": 22, "H": 30, "M": 26}.items():
        ws.column_dimensions[col].width = w


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: CO Attainment Criteria (master weightage config, matches the
#  reference workbook's validation table — direct 80% / indirect 20%).
# ═══════════════════════════════════════════════════════════════════════
def build_criteria_sheet(wb, category_weights: Dict[str, float]):
    ws = wb.create_sheet("CO Attainment Criteria")
    ws["A2"] = "INSTRUCTIONS: IF YOU WANT TO CHANGE THE WEIGHTAGE, ENTER THE CHANGED WEIGHTAGE IN THE ORANGE BOX ONLY"
    ws["A2"].font = _font(italic=True)

    ct = category_weights.get("Class Test", 0)
    quiz = category_weights.get("MCQ/Quiz", 0)
    ass = category_weights.get("Laboratory/Assignment/Activity", 0)
    moocs = category_weights.get("MOOCs", 0)
    msa = category_weights.get("Mid Semester Assessment", 0)
    esa = category_weights.get("End Semester Assessment", 0)
    ces = category_weights.get("Course Exit Survey", 0)
    ca_total = round(ct + quiz + ass + moocs, 3)
    direct_total = round(ca_total + msa + esa, 3)
    indirect_total = ces

    ws["B5"] = "Component"; ws["D5"] = "Weightage"
    for c in ("B5", "D5"):
        ws[c].font = WHITE_BOLD; ws[c].fill = BRIGHT_BLUE

    rows = [
        ("DIRECT", "Direct Attainment", direct_total, True),
        ("CA", "Continuous Assesement", ca_total, False),
        ("CT", "Class Test", ct, False),
        ("QUIZ", "MCQ/Quiz", quiz, False),
        ("LAB/ASS/ACT", "Laboratory/Assignment/Activity", ass, False),
        ("EXP L", "MOOCs", moocs, False),
        ("MSA", "Mid Semester Assessment", msa, False),
        ("ESA", "End Semester Assessment", esa, False),
        ("INDIRECT", "Indirect Attainment", indirect_total, True),
        ("CES", "Course Exit Survey", ces, False),
    ]
    for i, (code, label, val, is_group) in enumerate(rows):
        r = 6 + i
        code_cell = ws.cell(row=r, column=2, value=code)
        ws.cell(row=r, column=3, value=label)
        val_cell = ws.cell(row=r, column=4, value=val)
        val_cell.fill = YELLOW if is_group else LABEL_BLUE
        if is_group:
            code_cell.font = BOLD
            ws.cell(row=r, column=3).font = BOLD
            val_cell.font = BOLD

    ws.cell(row=11, column=8, value=direct_total).fill = GREEN_HDR
    ws.cell(row=11, column=9, value="DIRECT").font = BOLD
    ws.cell(row=11, column=18, value="INDIRECT").font = BOLD
    ws.cell(row=11, column=20, value=indirect_total).fill = GREEN_HDR

    border_range(ws, 5, 15, 2, 4)
    for c, w in {"A": 10, "B": 16, "C": 34, "D": 12}.items():
        ws.column_dimensions[c].width = w


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: one per CATEGORY (Class Tests, Quiz, Assignment, MOOCs, Mid-Sem,
#  End-Sem, CES) — same layout as the reference workbook: every sub-
#  instrument in the category (e.g. CT-1/CT-2/CT-3) is laid out as its own
#  column block, side by side with no gap, followed by a combined CO Wise
#  Marks/Percentage section computed by summing obtained/max marks for
#  each CO *across all blocks* for the same student (mirrors the
#  reference's SUMIFS-across-test-columns formula), and a single
#  attendance/scoring/Level summary per CO for the whole category.
# ═══════════════════════════════════════════════════════════════════════
def build_category_sheet(wb, category_label, members, targets, co_attainment_target, course_name):
    sheet_name = CATEGORY_SHEET_NAMES.get(category_label, category_label)[:31]
    ws = wb.create_sheet(title=sheet_name)

    total_weight = round(sum(m.get("weightage", 0) or 0 for m in members), 3)

    ws["D2"] = "CO TARGET"; ws["D2"].font = BOLD
    comparison_pct = targets[0]["comparision_percentage"] if targets else None
    ws["A3"] = "Threshold value for target:"; ws["C3"] = comparison_pct
    ws["A4"] = "SET CO TARGET"; ws["D4"] = co_attainment_target
    ws["H6"] = f"{sheet_name.upper()}  (weight {total_weight}% of final grade)"
    ws["H6"].font = BOLD

    if not members or not any(m.get("students") for m in members):
        msg = "No MOOCs component configured for this subject." if category_label == "MOOCs" \
            else "No students found for this category."
        ws.cell(row=8, column=1, value=msg)
        return {}, sheet_name

    multi = len(members) > 1
    r_banner, r_qno, r_co, r_marks = 6, 7, 8, 9

    ws.cell(row=r_qno, column=1,
            value="INTERNAL ASSESSMENT MARKS" if category_label == "Class Test" else category_label.upper())

    col = 5  # E
    block_meta: List[Dict[str, Any]] = []
    for idx, m in enumerate(members):
        columns = build_columns(m.get("students", []))
        block_start = col
        for i, c in enumerate(columns):
            ws.cell(row=r_qno, column=block_start + i, value=c["questionNo"])
            ws.cell(row=r_co, column=block_start + i, value=c["coCode"])
            ws.cell(row=r_marks, column=block_start + i, value=c["maxMarks"])
        total_col = block_start + len(columns)
        ws.cell(row=r_qno, column=total_col, value="TOTAL")
        if multi:
            label = f"TEST-{idx + 1}" if category_label == "Class Test" else short_exam_label(m.get("folder_name", ""), course_name)
            bc = ws.cell(row=r_banner, column=block_start, value=label)
            bc.font = BOLD
            if total_col - 1 > block_start:
                ws.merge_cells(start_row=r_banner, start_column=block_start, end_row=r_banner, end_column=total_col - 1)
        block_meta.append({"member": m, "columns": columns, "start": block_start, "total_col": total_col})
        col = total_col + 1

    last_total_col = col - 1

    ws.cell(row=r_qno, column=4, value="Q. No")
    ws.cell(row=r_co, column=4, value="COs")
    ws.cell(row=r_marks, column=1, value="Sl.No.")
    ws.cell(row=r_marks, column=2, value="Name")
    ws.cell(row=r_marks, column=3, value="Reg. No.")
    ws.cell(row=r_marks, column=4, value="MARKS")

    for rr in (r_qno, r_co, r_marks):
        for c in range(1, last_total_col + 1):
            cell = ws.cell(row=rr, column=c)
            cell.alignment = CENTER
            if c <= 3:
                cell.font = WHITE_BOLD; cell.fill = ROSTER_BLUE
            elif c == 4:
                cell.font = WHITE_BOLD; cell.fill = DARK_BLUE
            elif any(c == b["total_col"] for b in block_meta):
                cell.font = BOLD; cell.fill = YELLOW
            else:
                cell.font = BOLD; cell.fill = LIGHT_BLUE_DATA
    ws.row_dimensions[r_marks].height = 31.2
    for r in range(r_banner if multi else r_qno, r_marks + 1):
        ws.row_dimensions[r].height = 15.6

    all_co_codes = sorted({c["coCode"] for b in block_meta for c in b["columns"]})
    n_co = len(all_co_codes)

    co_marks_col0 = last_total_col + 2
    co_pct_col0 = co_marks_col0 + n_co + 1

    marks_title = ws.cell(row=r_qno, column=co_marks_col0, value="CO Wise Marks")
    marks_title.font = BOLD; marks_title.fill = LIGHT_BLUE_DATA; marks_title.alignment = CENTER
    if n_co > 1:
        ws.merge_cells(start_row=r_qno, start_column=co_marks_col0, end_row=r_qno, end_column=co_marks_col0 + n_co - 1)
    for i in range(1, n_co):
        ws.cell(row=r_qno, column=co_marks_col0 + i).fill = LIGHT_BLUE_DATA

    pct_title = ws.cell(row=r_qno, column=co_pct_col0, value="CO Wise Percentage")
    pct_title.font = WHITE_BOLD; pct_title.fill = DARK_NAVY; pct_title.alignment = CENTER
    if n_co > 1:
        ws.merge_cells(start_row=r_qno, start_column=co_pct_col0, end_row=r_qno, end_column=co_pct_col0 + n_co - 1)
    for i in range(1, n_co):
        ws.cell(row=r_qno, column=co_pct_col0 + i).fill = DARK_NAVY

    for i, co in enumerate(all_co_codes):
        hc = ws.cell(row=r_co, column=co_marks_col0 + i, value=co)
        hc.font = BOLD; hc.fill = LIGHT_BLUE_DATA; hc.alignment = CENTER
        hc = ws.cell(row=r_co, column=co_pct_col0 + i, value=co)
        hc.font = WHITE_BOLD; hc.fill = DARK_NAVY; hc.alignment = CENTER

    crit_col = co_pct_col0 + n_co + 2
    tgt_cell = ws.cell(row=r_co, column=crit_col, value="Target")
    tgt_cell.font = WHITE_BOLD; tgt_cell.fill = GREEN_HDR; tgt_cell.alignment = CENTER
    for i, (label, level) in enumerate(criteria_labels()):
        r = r_marks + i
        ws.cell(row=r, column=crit_col, value=label).fill = GREEN_HDR
        ws.cell(row=r, column=crit_col + 4, value=comparison_pct).fill = GREEN_HDR
        ws.cell(row=r, column=crit_col + 5, value="marks").fill = GREEN_HDR
        ws.cell(row=r, column=crit_col + 7, value=level).fill = GREEN_HDR

    # Master roster — union of students across all blocks (all sub-instruments
    # in a category share the same class roster), keyed by filename so marks
    # from different blocks line up on the same student row.
    roster: Dict[str, Dict[str, Any]] = {}
    for b in block_meta:
        for s in b["member"].get("students", []):
            key = s.get("filename") or s.get("student_name")
            if key and key not in roster:
                roster[key] = {"student_name": s.get("student_name"), "filename": s.get("filename")}
    roster_list = sorted(roster.values(), key=lambda s: (s["student_name"] or s["filename"] or "").lower())

    # Constant per-CO totals/source-columns (same for every student row) — used
    # to build real Excel formulas below instead of static computed values.
    co_max_total: Dict[str, float] = {co: 0.0 for co in all_co_codes}
    co_source_cols: Dict[str, List[int]] = {co: [] for co in all_co_codes}
    for b in block_meta:
        for i, c in enumerate(b["columns"]):
            co_max_total[c["coCode"]] += c["maxMarks"] or 0
            co_source_cols[c["coCode"]].append(b["start"] + i)

    data_start = r_marks + 1
    co_student_pcts: Dict[str, List[Optional[float]]] = {co: [] for co in all_co_codes}

    for si, ros in enumerate(roster_list):
        row = data_start + si
        key = ros.get("filename") or ros.get("student_name")
        ws.cell(row=row, column=1, value=si + 1)
        ws.cell(row=row, column=2, value=ros.get("student_name") or ros.get("filename") or f"Student {si + 1}")
        ws.cell(row=row, column=3, value=ros.get("filename", ""))

        co_obtained_total: Dict[str, float] = {co: 0.0 for co in all_co_codes}

        for b in block_meta:
            student = next(
                (s for s in b["member"].get("students", [])
                 if (s.get("filename") or s.get("student_name")) == key), None,
            )
            lookup = {}
            if student:
                for q in student.get("questions", []):
                    for co in q.get("cos", []):
                        lookup[f"{q['question_no']}|{co['co_code']}"] = co.get("obtained_marks")
            for i, c in enumerate(b["columns"]):
                val = lookup.get(f"{c['questionNo']}|{c['coCode']}")
                ws.cell(row=row, column=b["start"] + i, value=val if val is not None else 0)
                co_obtained_total[c["coCode"]] += val or 0
            block_start_letter = get_column_letter(b["start"])
            block_end_letter = get_column_letter(b["total_col"] - 1)
            ws.cell(row=row, column=b["total_col"],
                    value=f"=SUM({block_start_letter}{row}:{block_end_letter}{row})").font = BOLD

        for i, co in enumerate(all_co_codes):
            marks_col = co_marks_col0 + i
            pct_col = co_pct_col0 + i
            src_refs = [f"{get_column_letter(cidx)}{row}" for cidx in co_source_cols[co]]
            ws.cell(row=row, column=marks_col, value=("=" + "+".join(src_refs)) if src_refs else 0)
            marks_letter = get_column_letter(marks_col)
            if co_max_total[co]:
                ws.cell(row=row, column=pct_col, value=f"=ROUND({marks_letter}{row}/{co_max_total[co]}*100,2)")
            else:
                ws.cell(row=row, column=pct_col, value="N/A")
            pct = calc_pct(co_obtained_total[co], co_max_total[co]) if co_max_total[co] else None
            # Matches the reference's COUNTIF(">=0") attendance formula: every
            # student in the roster counts, missing data treated as 0%, not
            # excluded.
            co_student_pcts[co].append(pct if pct is not None else 0.0)

    lab_row = data_start + len(roster_list) + 2
    ws.cell(row=lab_row, column=crit_col, value="CO'S").font = WHITE_BOLD
    ws.cell(row=lab_row, column=crit_col).fill = GREEN_HDR
    for i, co in enumerate(all_co_codes):
        c = ws.cell(row=lab_row, column=crit_col + 6 + i, value=co)
        c.font = WHITE_BOLD; c.fill = GREEN_HDR; c.alignment = CENTER

    row_labels = ["No. of students Attended", "# Students scoring>=", "% Students scoring>=", "Attainment"]
    for li, lab in enumerate(row_labels):
        ws.cell(row=lab_row + 1 + li, column=crit_col, value=lab).font = BOLD

    per_co: Dict[str, Any] = {}
    for i, co in enumerate(all_co_codes):
        pcts = co_student_pcts[co]
        attended = len(pcts)
        above = sum(1 for p in pcts if comparison_pct is not None and p >= comparison_pct) if attended else 0
        ach_pct = round(above / attended * 100, 2) if attended else 0
        level = level_for_pct(ach_pct, targets) if attended else 0

        c = crit_col + 6 + i
        col_letter = get_column_letter(c)
        pct_col_letter = get_column_letter(co_pct_col0 + i)
        pct_range = (
            f"{pct_col_letter}{data_start}:{pct_col_letter}{data_start + len(roster_list) - 1}"
            if roster_list else None
        )
        attended_cell = f"{col_letter}{lab_row + 1}"
        above_cell = f"{col_letter}{lab_row + 2}"
        ach_cell = f"{col_letter}{lab_row + 3}"

        if pct_range:
            ws.cell(row=lab_row + 1, column=c, value=f'=COUNTIF({pct_range},">=0")')
            ws.cell(row=lab_row + 2, column=c, value=f'=COUNTIF({pct_range},">="&$C$3)')
        else:
            ws.cell(row=lab_row + 1, column=c, value=0)
            ws.cell(row=lab_row + 2, column=c, value=0)
        ws.cell(row=lab_row + 3, column=c,
                value=f'=IF({attended_cell}=0,0,ROUND({above_cell}/{attended_cell}*100,2))')

        level_expr = _level_if_formula(ach_cell, targets)
        lvl_cell = ws.cell(row=lab_row + 4, column=c, value=f'=IF({attended_cell}=0,"N/A",{level_expr})')
        lvl_cell.font = BOLD
        if attended:
            lvl_cell.fill = GREY
        per_co[co] = {
            "level": level if attended else None, "achievementPct": ach_pct if attended else None,
            "weight": total_weight, "level_cell": f"'{sheet_name}'!{col_letter}{lab_row + 4}",
        }

    border_range(ws, r_qno, lab_row + 4, 1, crit_col + 6 + n_co - 1)

    for c, w in {"A": 8, "B": 25.2, "C": 14, "D": 16.5}.items():
        ws.column_dimensions[c].width = w

    return per_co, sheet_name


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: Final CO Attainment
# ═══════════════════════════════════════════════════════════════════════
def build_final_co_attainment_sheet(wb, meta, category_reports, all_cos, co_attainment_target,
                                     co_po_matrix_obj) -> Dict[str, float]:
    ws = wb.create_sheet("Final CO Attainment")
    pos = po_codes_from_matrix(co_po_matrix_obj)
    grid_width = max(1 + len(pos), 6)  # section banners span at least this many columns

    ws["A1"] = "COURSE"; ws["B1"] = meta.get("course", "")
    ws["A2"] = "COURSE CODE"; ws["B2"] = meta.get("course_code", "")
    ws["A3"] = "FACULTY"; ws["B3"] = meta.get("faculty_display", "")
    ws["A4"] = "BRANCH/SECTION"; ws["B4"] = meta.get("branch", "")
    ws["L1"] = "ACADEMIC YEAR:"; ws["M1"] = meta.get("academic_year", "")
    ws["L2"] = "SEMESTER:"; ws["M2"] = str(meta.get("semester", ""))
    for c in ("A1", "A2", "A3", "A4", "B1", "B2", "B3", "B4"):
        ws[c].font = BOLD; ws[c].fill = LABEL_ORANGE
    for c in ("L1", "L2", "M1", "M2"):
        ws[c].font = BOLD; ws[c].fill = LABEL_ORANGE
    border_range(ws, 1, 4, 1, 2)
    border_range(ws, 1, 2, 12, 13)

    banner = ws.cell(row=6, column=1, value="CO Attainments")
    banner.font = WHITE_BOLD; banner.fill = BRIGHT_BLUE; banner.alignment = CENTER
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=grid_width)

    hdr_row = 7
    for i, po in enumerate(pos):
        cell = ws.cell(row=hdr_row, column=2 + i, value=po)
        cell.font = WHITE_BOLD; cell.fill = BRIGHT_BLUE; cell.alignment = CENTER
    for r, co in enumerate(all_cos):
        row = hdr_row + 1 + r
        ws.cell(row=row, column=1, value=co).font = BOLD
        ws.cell(row=row, column=1).fill = YELLOW
        for i, po in enumerate(pos):
            val = (co_po_matrix_obj.get(co) or {}).get(po)
            cell = ws.cell(row=row, column=2 + i, value=val if val is not None else "")
            cell.alignment = CENTER; cell.fill = LABEL_BLUE
    avg_row = hdr_row + 1 + len(all_cos)
    ws.cell(row=avg_row, column=1, value="Average").font = BOLD
    ws.cell(row=avg_row, column=1).fill = YELLOW
    for i, po in enumerate(pos):
        col_letter = get_column_letter(2 + i)
        cell = ws.cell(
            row=avg_row, column=2 + i,
            value=f"=IFERROR(AVERAGE({col_letter}{hdr_row + 1}:{col_letter}{avg_row - 1}),\"\")",
        )
        cell.font = BOLD; cell.fill = YELLOW; cell.alignment = CENTER
    if pos:
        border_range(ws, hdr_row, avg_row, 1, 1 + len(pos))

    # Level/Target legend columns must clear BOTH the CO columns here AND the
    # (possibly much wider) CO-PO matrix above — otherwise they land on top
    # of PO9/PO10/etc's columns and read as garbled overlap.
    legend_col = max(2 + len(all_cos), 2 + len(pos)) + 2
    r16 = avg_row + 3
    ws.cell(row=r16, column=1, value="Cos").font = BOLD
    for i, co in enumerate(all_cos):
        ws.cell(row=r16, column=2 + i, value=co).font = BOLD
    ws.cell(row=r16, column=legend_col, value="Level").font = WHITE_BOLD
    ws.cell(row=r16, column=legend_col + 1, value="Target").font = WHITE_BOLD
    ws.merge_cells(start_row=r16, start_column=legend_col + 1, end_row=r16, end_column=legend_col + 2)
    for c in range(1, 2 + len(all_cos)):
        ws.cell(row=r16, column=c).fill = YELLOW
    ws.cell(row=r16, column=legend_col).fill = BRIGHT_BLUE
    ws.cell(row=r16, column=legend_col + 1).fill = BRIGHT_BLUE

    ws.cell(row=r16 + 1, column=1, value="Attainment Level  as per set rules")
    for i, co in enumerate(all_cos):
        ws.cell(row=r16 + 1, column=2 + i, value=co_attainment_target)
    comparison_pct = None
    for i, (label, level) in enumerate(criteria_labels()):
        row = r16 + 1 + i
        ws.cell(row=row, column=legend_col, value=level).alignment = CENTER
        ws.cell(row=row, column=legend_col + 1, value=label)
        ws.merge_cells(start_row=row, start_column=legend_col + 1, end_row=row, end_column=legend_col + 2)
    border_range(ws, r16, r16 + 3, 1, 1 + len(all_cos))
    border_range(ws, r16, r16 + 3, legend_col, legend_col + 2)

    row_co_attain = r16 + 4
    ws.cell(row=row_co_attain, column=1, value="CO attainment").font = BOLD
    final_scores: Dict[str, float] = {}
    border_range(ws, row_co_attain, row_co_attain, 1, 1 + len(all_cos))

    # Group category sheets (Class Tests/Quiz/Assignment/MOOCs/MSA/ESA) into
    # Direct vs Indirect (CES), matching the reference workbook's Direct
    # Attainment (subtotal) + CES + Final Attainment layout. Each category
    # already carries a single, correctly-combined Level (computed in
    # build_category_sheet from marks summed across all its sub-instruments)
    # — no re-averaging needed here, just weight% * level per category.
    direct_categories: Dict[str, Dict[str, Any]] = {
        r["category"]: r for r in category_reports if not r["is_indirect"]
    }
    indirect_categories: Dict[str, Dict[str, Any]] = {
        r["category"]: r for r in category_reports if r["is_indirect"]
    }

    known_order = [label for _, label in _CATEGORY_PREFIXES]
    direct_order = sorted(direct_categories, key=lambda c: known_order.index(c) if c in known_order else 999)
    indirect_order = sorted(indirect_categories, key=lambda c: known_order.index(c) if c in known_order else 999)

    tbl_r = row_co_attain + 3
    title = ws.cell(row=tbl_r, column=1, value="(%) of Contribution — Direct + Indirect Attainment")
    title.font = WHITE_BOLD; title.fill = BRIGHT_BLUE; title.alignment = CENTER

    row24 = tbl_r + 1
    row25 = tbl_r + 2
    row26 = tbl_r + 3
    col = 2
    cat_cols: Dict[Any, int] = {}

    def place_category(cat_name, rep, header_label):
        nonlocal col
        weight = rep["weight"]
        ws.cell(row=row24, column=col, value=header_label)
        ws.cell(row=row24, column=col + 1, value=weight)
        ws.cell(row=row25, column=col + 1, value=weight / 100 if weight else 0)
        ws.cell(row=row26, column=col, value=f"{header_label} Level")
        ws.cell(row=row26, column=col + 1, value=f"x{weight}%")
        cat_cols[cat_name] = col
        col += 2

    for cat in direct_order:
        place_category(("D", cat), direct_categories[cat], cat)
    direct_subtotal_col = None
    if direct_order:
        direct_subtotal_col = col
        ws.cell(row=row24, column=col, value="Direct Attainment").font = BOLD
        ws.cell(row=row26, column=col, value="Direct").font = BOLD
        col += 1
    for cat in indirect_order:
        place_category(("I", cat), indirect_categories[cat], cat)
    final_col = col
    ws.cell(row=row24, column=final_col, value="Final Attainment").font = BOLD
    ws.cell(row=row26, column=final_col, value="Final").font = BOLD

    ws.merge_cells(start_row=tbl_r, start_column=1, end_row=tbl_r, end_column=final_col)
    ws.cell(row=row26, column=1, value="Course Outcomes").font = BOLD
    for c in range(1, final_col + 1):
        cell = ws.cell(row=row26, column=c)
        if cell.font is None or not cell.font.bold:
            cell.font = WHITE_BOLD
        cell.fill = BRIGHT_BLUE
        if ws.cell(row=row24, column=c).value is not None:
            ws.cell(row=row24, column=c).font = WHITE_BOLD
            ws.cell(row=row24, column=c).fill = BRIGHT_BLUE

    for i, co in enumerate(all_cos):
        row = row26 + 1 + i
        ws.cell(row=row, column=1, value=co).font = BOLD

        direct_contrib_refs: List[str] = []
        for cat in direct_order:
            c = cat_cols[("D", cat)]
            rep = direct_categories[cat]
            level_ref = rep["per_co"].get(co, {}).get("level_cell")
            col_letter = get_column_letter(c)
            weight_col_letter = get_column_letter(c + 1)
            ws.cell(row=row, column=c, value=f"=IF(ISNUMBER({level_ref}),{level_ref},0)" if level_ref else 0)
            ws.cell(row=row, column=c + 1, value=f"={col_letter}{row}*{weight_col_letter}${row25}")
            direct_contrib_refs.append(f"{weight_col_letter}{row}")
        direct_total_ref = None
        if direct_subtotal_col:
            subtotal_letter = get_column_letter(direct_subtotal_col)
            ws.cell(row=row, column=direct_subtotal_col,
                    value=("=" + "+".join(direct_contrib_refs)) if direct_contrib_refs else 0).font = BOLD
            direct_total_ref = f"{subtotal_letter}{row}"

        indirect_contrib_refs: List[str] = []
        for cat in indirect_order:
            c = cat_cols[("I", cat)]
            rep = indirect_categories[cat]
            level_ref = rep["per_co"].get(co, {}).get("level_cell")
            col_letter = get_column_letter(c)
            weight_col_letter = get_column_letter(c + 1)
            ws.cell(row=row, column=c, value=f"=IF(ISNUMBER({level_ref}),{level_ref},0)" if level_ref else 0)
            ws.cell(row=row, column=c + 1, value=f"={col_letter}{row}*{weight_col_letter}${row25}")
            indirect_contrib_refs.append(f"{weight_col_letter}{row}")

        final_refs = ([direct_total_ref] if direct_total_ref else direct_contrib_refs) + indirect_contrib_refs
        ws.cell(row=row, column=final_col, value=("=" + "+".join(final_refs)) if final_refs else 0).font = BOLD
        final_col_letter = get_column_letter(final_col)
        ws.cell(row=row_co_attain, column=2 + i, value=f"={final_col_letter}{row}").alignment = CENTER

        # Python-side numeric mirror — needed for the verdict formula's target
        # comparison text below and for Final PO attainment's attained/not-
        # attained bucketing (openpyxl can't evaluate formulas itself).
        direct_total = sum(
            (direct_categories[cat]["per_co"].get(co, {}).get("level") or 0) * direct_categories[cat]["weight"] / 100
            for cat in direct_order
        )
        indirect_total = sum(
            (indirect_categories[cat]["per_co"].get(co, {}).get("level") or 0) * indirect_categories[cat]["weight"] / 100
            for cat in indirect_order
        )
        final_scores[co] = round(direct_total + indirect_total, 3)
    border_range(ws, row26, row26 + len(all_cos), 1, final_col)

    # "Action Planned" gets its writing room from a per-row MERGE (C:F) rather
    # than forcing column C's width — column C is shared with the CO2 legend
    # column above, and widening it there threw off that row's alignment.
    action_end_col = 6
    verdict_r = row26 + 2 + len(all_cos) + 2
    ws.cell(row=verdict_r, column=1, value="CO ATTAINMENT").font = BOLD
    ws.cell(row=verdict_r, column=3, value="Action Planned for the next semester").font = BOLD
    ws.merge_cells(start_row=verdict_r, start_column=3, end_row=verdict_r, end_column=action_end_col)
    for c in (1, 2, 3):
        ws.cell(row=verdict_r, column=c).fill = BRIGHT_BLUE
        ws.cell(row=verdict_r, column=c).font = WHITE_BOLD
    verdict_cell_refs: Dict[str, str] = {}
    for i, co in enumerate(all_cos):
        row = verdict_r + 1 + i
        ws.cell(row=row, column=1, value=co)
        attain_ref = f"{get_column_letter(2 + i)}{row_co_attain}"
        target_ref = f"{get_column_letter(2 + i)}{r16 + 1}"
        rc = ws.cell(row=row, column=2, value=f'=IF({attain_ref}>={target_ref},"Attained","Not Attained")')
        rc.fill = GREY
        verdict_cell_refs[co] = f"'Final CO Attainment'!B{row}"
        ws.cell(row=row, column=3, value="")  # left for faculty to fill in
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=action_end_col)
    border_range(ws, verdict_r, verdict_r + len(all_cos), 1, action_end_col)

    autofit_columns(ws, min_width=10, max_width=20)
    ws.column_dimensions["A"].width = 36     # long row labels ("Attainment Level as per set rules")

    return final_scores, verdict_cell_refs


# ═══════════════════════════════════════════════════════════════════════
#  SHEET: Final PO attainment
# ═══════════════════════════════════════════════════════════════════════
def build_final_po_attainment_sheet(wb, all_cos, co_po_matrix_obj, final_scores, verdict_cell_refs, co_attainment_target):
    ws = wb.create_sheet("Final PO attainment")
    ws["A1"] = "PO Attainment"; ws["A1"].font = _font(bold=True, size=13)

    pos = po_codes_from_matrix(co_po_matrix_obj)
    if not pos or not all_cos:
        ws["A3"] = "No CO-PO matrix has been entered for this subject yet."
        return

    ws["A2"] = "COs"; ws["B2"] = "CO Results"
    for i, po in enumerate(pos):
        ws.cell(row=2, column=3 + i, value=po)
    for c in range(1, 3 + len(pos)):
        cell = ws.cell(row=2, column=c)
        cell.font = WHITE_BOLD; cell.fill = BRIGHT_BLUE

    for r, co in enumerate(all_cos):
        row = 3 + r
        ws.cell(row=row, column=1, value=co).font = BOLD
        ref = verdict_cell_refs.get(co)
        rc = ws.cell(row=row, column=2, value=f"={ref}" if ref else "")
        rc.fill = GREY
        for i, po in enumerate(pos):
            val = (co_po_matrix_obj.get(co) or {}).get(po)
            ws.cell(row=row, column=3 + i, value=val if val is not None else "").alignment = CENTER

    last_co_row = 2 + len(all_cos)
    pct_row = 3 + len(all_cos)
    lvl_row = pct_row + 1
    ws.cell(row=pct_row, column=1, value="PO Attainment Percentage").font = BOLD
    ws.cell(row=lvl_row, column=1, value="PO Attained").font = BOLD
    status_range = f"$B$3:$B${last_co_row}"
    for i, po in enumerate(pos):
        col_letter = get_column_letter(3 + i)
        data_range = f"{col_letter}3:{col_letter}{last_co_row}"
        pct_cell = ws.cell(
            row=pct_row, column=3 + i,
            value=f'=ROUND((SUMIF({status_range},"Attained",{data_range})/SUM({data_range}))*100,2)',
        )
        pct_cell.alignment = CENTER
        lvl_cell = ws.cell(row=lvl_row, column=3 + i, value=f"=ROUND(({col_letter}{pct_row}/100)*3,2)")
        lvl_cell.alignment = CENTER
    border_range(ws, 2, lvl_row, 1, 2 + len(pos))

    autofit_columns(ws, min_width=10, max_width=28)


# ═══════════════════════════════════════════════════════════════════════
#  ENTRY POINTS (called from app/api/routers/subject_results.py)
# ═══════════════════════════════════════════════════════════════════════
async def fetch_co_report_data(db: AsyncIOMotorDatabase, subject_id: str) -> Dict[str, Any]:
    """Async DB fetch + the same guard conditions the router turns into 400/404s.
    Raises CoReportError on failure; otherwise returns the data build_co_report_workbook needs."""
    if not subject_id:
        raise CoReportError(400, {"error": "subjectId is required"})

    exams = await fetch_exams(db, subject_id)
    targets, co_att_tgt = await fetch_targets(db, subject_id)
    co_po_matrix_obj, co_po_matrix_arr = await fetch_co_po(db, subject_id)
    meta = await fetch_subject_meta(db, subject_id)

    if not exams:
        raise CoReportError(404, {"error": "No evaluations found for this subject"})

    total_students = sum(len(e["students"]) for e in exams)
    if total_students == 0:
        sample = [s async for s in db["answerDetails"].find().limit(2)]
        evals = [e async for e in db["newsavedDocs"].find({"subject_id": ObjectId(subject_id)})]
        raise CoReportError(404, {
            "error": "No answer sheets found — check evaluation_id field.",
            "first_eval_id": str(evals[0]["_id"]) if evals else None,
            "sample_keys": [list(s.keys()) for s in sample],
        })

    return {
        "exams": exams, "targets": targets, "co_attainment_target": co_att_tgt,
        "co_po_matrix_obj": co_po_matrix_obj, "co_po_matrix_arr": co_po_matrix_arr,
        "meta": meta,
    }


def build_co_report_workbook(data: Dict[str, Any]) -> bytes:
    """Pure/sync — run via asyncio.to_thread from the router."""
    exams = data["exams"]
    targets = data["targets"]
    co_att_tgt = data["co_attainment_target"]
    co_po_matrix_obj = data["co_po_matrix_obj"]
    meta = data["meta"]

    all_cos = sorted({
        co["co_code"] for exam in exams for student in exam.get("students", [])
        for q in student.get("questions", []) for co in q.get("cos", [])
    })

    wb = Workbook()
    wb.remove(wb.active)

    build_start_sheet(wb, meta, targets, co_att_tgt, co_po_matrix_obj, all_cos)

    # Group exams into reference-style categories (Class Test/Quiz/Assignment/
    # MOOCs/MSA/ESA/CES) — each category becomes ONE sheet with its sub-
    # instruments side by side, matching the reference workbook exactly
    # instead of one sheet per individual exam.
    categories: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for exam in exams:
        cat = exam_category(exam)
        if cat not in categories:
            categories[cat] = {"members": [], "is_indirect": bool(exam.get("is_course_exit_summary"))}
            order.append(cat)
        categories[cat]["members"].append(exam)

    known_order = [label for _, label in _CATEGORY_PREFIXES]
    if "MOOCs" not in categories:
        categories["MOOCs"] = {"members": [], "is_indirect": False}
        order.append("MOOCs")
    order = sorted(order, key=lambda c: known_order.index(c) if c in known_order else 999)

    category_weights = {
        cat: round(sum(m.get("weightage", 0) or 0 for m in categories[cat]["members"]), 3) for cat in order
    }
    build_criteria_sheet(wb, category_weights)

    category_reports = []
    for cat in order:
        info = categories[cat]
        per_co, sheet_name = build_category_sheet(wb, cat, info["members"], targets, co_att_tgt, meta.get("course", ""))
        category_reports.append({
            "category": cat, "sheet_name": sheet_name, "weight": category_weights[cat],
            "is_indirect": info["is_indirect"], "per_co": per_co,
        })

    final_scores, verdict_cell_refs = build_final_co_attainment_sheet(
        wb, meta, category_reports, all_cos, co_att_tgt, co_po_matrix_obj
    )
    build_final_po_attainment_sheet(wb, all_cos, co_po_matrix_obj, final_scores, verdict_cell_refs, co_att_tgt)

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_format.defaultRowHeight = 15.6

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
